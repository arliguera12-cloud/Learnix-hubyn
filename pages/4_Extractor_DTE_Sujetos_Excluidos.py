import functools
import streamlit as st
import pdfplumber
import pandas as pd
import re
import json
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from styles import DARK_PRO_CSS
from utils.concurrent_processor import leer_y_procesar_lote
from utils.pdf_utils import (
    safe_str,
    normalizar_unicode,
    limpiar_monto,
    extraer_y_formatear_fecha,
    extraer_texto_pdf,
)
from utils.gemini_utils import (
    gemini_disponible,
    gemini_ultimo_error,
    procesar_dte_con_gemini,
)
from utils.gemini_vision import (
    extraer_dte_con_vision,
    vision_disponible,
    vision_ultimo_error,
)
from utils.qa_utils import (
    mostrar_banner_qa,
    mostrar_indicador_vision,
    validar_montos_sujetos,
)

# ─────────────────────────────────────────────
# 1. PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(page_title="Extractor DTE · Sujetos Excluidos", layout="wide", page_icon="⚖️")

# ─────────────────────────────────────────────
# 2. ESTILOS
# ─────────────────────────────────────────────
st.markdown(DARK_PRO_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 3. SEGURIDAD — Multi-tenant SaaS
# ─────────────────────────────────────────────
from utils.auth_guard import check_auth
check_auth()

if not st.session_state.get("cliente_activo"):
    st.warning("⚠️ Debes seleccionar un Cliente Activo en el Dashboard antes de extraer Sujetos Excluidos.")
    st.stop()

cliente = st.session_state.cliente_activo

# ─────────────────────────────────────────────
# 4. FUNCIONES AUXILIARES
# ─────────────────────────────────────────────






def limpiar_nit(raw: str) -> str:
    return re.sub(r'[^0-9]', '', raw)


def extraer_nombre_receptor(texto: str) -> str:
    """
    Los DTE-14 de FULLCHEM tienen EMISOR y RECEPTOR en la misma línea:
      "Nombre o razón social: INDUSTRIAS FULLCHEM... Nombre o razón social: CARLOS ENRIQUE SASSO LEMUS"
    
    La clave es tomar la SEGUNDA ocurrencia de "Nombre o razón social:" en la línea.
    También manejamos el caso donde están en líneas separadas (otros emisores).
    """
    # ── Estrategia 1: misma línea — buscar segunda ocurrencia del patrón en una sola línea ──
    # Patrón: dos "Nombre o razón social:" en la misma línea
    m = re.search(
        r"[Nn]ombre\s+o\s+raz[oó]n\s+social\s*:\s*.+?"   # primera (emisor)
        r"[Nn]ombre\s+o\s+raz[oó]n\s+social\s*:\s*"       # segunda etiqueta (receptor)
        r"([A-ZÁÉÍÓÚÜÑ][A-ZÁÉÍÓÚÜÑa-záéíóúüñ ,.\-]+)",   # nombre del receptor
        texto, re.I
    )
    if m:
        nombre = re.sub(r'\s+', ' ', m.group(1)).strip()
        # Cortar si llega a NIT: o DUI: u otro campo
        nombre = re.split(r'\s+(?:NIT|DUI|N[úu]mero|Direcci[oó]n|Correo|Tel[eé]fono)\s*[:\-]', nombre, flags=re.I)[0]
        nombre = nombre.strip()
        if 3 < len(nombre) <= 65:
            return nombre.upper()

    # ── Estrategia 2: líneas separadas — tomar el último "Nombre o razón social:" ──
    todas = re.findall(
        r"[Nn]ombre\s+o\s+raz[oó]n\s+social\s*[:\-]?\s*([^\n]+)",
        texto, re.I
    )
    if len(todas) >= 2:
        nombre = re.sub(r'\s+', ' ', todas[-1]).strip()
        nombre = re.split(r'\s+(?:NIT|DUI|N[úu]mero|Direcci[oó]n|Correo|Tel[eé]fono)\s*[:\-]', nombre, flags=re.I)[0]
        nombre = nombre.strip()
        if 3 < len(nombre) <= 65:
            return nombre.upper()
    elif len(todas) == 1:
        nombre = re.sub(r'\s+', ' ', todas[0]).strip()
        if 3 < len(nombre) <= 65:
            return nombre.upper()

    return "⚠️ REVISAR NOMBRE"


def extraer_nit_receptor(texto: str, nit_cliente: str) -> str:
    """
    Similar al nombre: en DTE-14 FULLCHEM el NIT del receptor está en la MISMA línea
    que el NIT del emisor: "NIT: 0614-270815-107-7  NIT: 07515732-7"
    Tomamos el último NIT que NO sea el del cliente/emisor.
    """
    # Buscar todos los NITs/DUIs del texto
    patron = (
        r"\b\d{4}\s*-\s*\d{6}\s*-\s*\d{3}\s*-\s*\d\b"   # NIT empresa 0614-270815-107-7
        r"|\b\d{8}\s*-\s*\d\b"                              # DUI 07515732-7
        r"|\b\d{14}\b"                                       # NIT sin guiones
        r"|\b\d{9}\b"                                        # DUI sin guión
    )
    ids_raw   = re.findall(patron, texto)
    ids_limp  = list(dict.fromkeys(limpiar_nit(n) for n in ids_raw))
    candidatos = [n for n in ids_limp if n != nit_cliente and len(n) >= 8]

    # El receptor suele ser el ÚLTIMO NIT distinto al emisor (que va primero)
    if candidatos:
        return candidatos[-1]
    return ""


def extraer_sello_dte14(texto: str) -> str:
    """Extrae el Sello de Recepción del DTE-14."""
    # Intento 1: etiqueta explícita
    m = re.search(
        r"[Ss]ello\s+(?:de\s+)?[Rr]ecepci[oó]n\s*[:\-]?\s*([A-Z0-9]{20,50})",
        texto, re.I
    )
    if m:
        return m.group(1).strip()[:40]

    # Intento 2: cadena año + 36 chars alfanuméricos en texto sin espacios
    t_ns = re.sub(r'\s+', '', texto).upper()
    m2 = re.search(r'(20[2-3]\d[A-Z0-9]{36})', t_ns)
    if m2:
        return m2.group(1)

    # Intento 3: "SELLO" seguido de la cadena en texto compacto
    m3 = re.search(r'SELLO[A-Z]*:?([A-Z0-9]{30,50})', t_ns)
    if m3:
        return m3.group(1)[:40]

    # Intento 4: línea que sea exactamente un sello (año + alfanumérico, sin guiones)
    for linea in texto.splitlines():
        linea_s = linea.strip()
        mc = re.match(r'^(20[2-3]\d[A-Z0-9]{26,})$', linea_s, re.I)
        if mc:
            candidato = mc.group(1).upper()
            if '-' not in candidato:
                return candidato

    return ""


def extraer_sujetos_nativo(file_bytes: bytes, cliente_activo: dict) -> dict:
    if not file_bytes or len(file_bytes) < 512:
        return {"error": "Archivo vacío o corrupto."}

    # ── Vision-First: extraer con IA antes de pdfplumber ─────────────────────
    _nit_cliente_ctx = limpiar_nit(cliente_activo.get('nit', ''))
    _nom_cliente_ctx = cliente_activo.get('nombre', '')

    gemini_correcciones: list[str] = []
    _vision_campos: dict  = {}
    _vision_alertas: list = []
    _vision_audit: dict   = {}

    if vision_disponible():
        _vision_campos, _vision_alertas, _vision_audit = extraer_dte_con_vision(
            file_bytes,
            "sujetos_excluidos",
            {"nit": _nit_cliente_ctx, "nombre": _nom_cliente_ctx},
        )
        gemini_correcciones = [
            f"Vision: {a}" for a in _vision_alertas
        ] if _vision_alertas else (
            [f"Vision extrajo {len(_vision_campos)} campo(s)"]
            if _vision_campos else []
        )

    try:
        texto_lineal, texto_visual = extraer_texto_pdf(file_bytes)
        texto_completo = texto_lineal + "\n" + texto_visual

        if len(texto_completo.strip()) < 50 and not _vision_campos.get("num_control"):
            return {"error": "PDF de imagen — sin texto extraíble."}

        t_clean = re.sub(r'[ \t]+', ' ', texto_completo)
        t_no_sp = re.sub(r'\s+', '', t_clean).upper()

        # ── Tipo DTE ──
        m_ctrl = re.search(r"(DTE-[0-9O]{2}-[A-Z0-9]{1,20}-\d{9,18})", t_no_sp)
        tipo   = "14"
        if m_ctrl:
            ctrl   = m_ctrl.group(1).replace("O", "0")
            m_tipo = re.search(r"DTE-(\d{2})", ctrl)
            if m_tipo:
                tipo = m_tipo.group(1)

        if tipo != "14":
            return {"error_tipo": f"Documento DTE-{tipo}. Solo se admiten DTE-14 (Sujetos Excluidos)."}

        nit_cliente = limpiar_nit(cliente_activo.get('nit', ''))

        # ── UUID / Código de Generación ──
        gen = ""
        m_uuid = re.search(
            r"([A-F0-9]{8}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{12})",
            t_no_sp
        )
        if m_uuid:
            raw = m_uuid.group(1).replace("-", "")
            gen = f"{raw[:8]}-{raw[8:12]}-{raw[12:16]}-{raw[16:20]}-{raw[20:]}"

        # ── Número de Control ──
        num_control = ""
        m_nc = re.search(r"(DTE-14-[A-Z0-9]+-\d+)", t_no_sp)
        if m_nc:
            num_control = m_nc.group(1)

        # ── Sello de Recepción ──
        sello = extraer_sello_dte14(t_clean)

        fecha = extraer_y_formatear_fecha(t_clean)

        # ── Nombre del receptor ──
        nom_sujeto = extraer_nombre_receptor(t_clean)

        # ── NIT/DUI del receptor ──
        id_sujeto = extraer_nit_receptor(t_clean, nit_cliente)

        # ── Montos: Base (Sumatoria ventas / Sub-Total) y Retención Renta 10% ──
        base = 0.0
        ret  = 0.0

        # Retención Renta — etiqueta explícita
        m_ret_renta = re.search(
            r"[Rr]etenci[oó]n\s+[Rr]enta\s*[:\-]?\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?)",
            t_clean
        )
        if m_ret_renta:
            ret = limpiar_monto(m_ret_renta.group(1))

        # Sub-Total o Sumatoria de ventas
        m_base = re.search(
            r"(?:[Ss]umatoria\s+de\s+[Vv]entas|[Ss]ub[-\s]?[Tt]otal)\s*[:\-]?\s*"
            r"(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?)",
            t_clean
        )
        if m_base:
            base = limpiar_monto(m_base.group(1))

        # Fallback: relación matemática 10%
        if base == 0.0 and ret > 0:
            base = round(ret * 10, 2)

        if base == 0.0:
            montos_raw = re.findall(
                r"(?:US\$?|\$)?\s*(\d{1,3}(?:[.,]\d{3})*[.,]\d{1,2})",
                t_clean
            )
            valores = sorted(
                list({limpiar_monto(m) for m in montos_raw if limpiar_monto(m) > 0}),
                reverse=True
            )
            for v in valores:
                ret_calc = round(v * 0.10, 2)
                if any(abs(r - ret_calc) <= 0.05 for r in valores if r < v):
                    base = v
                    if ret == 0:
                        ret = ret_calc
                    break

        if base > 0 and ret == 0:
            ret = round(base * 0.10, 2)

        # ── Aplicar Vision con prioridad sobre regex ──────────────────────────
        if _vision_campos.get("fecha"):
            fecha = _vision_campos["fecha"]
        if _vision_campos.get("nom_sujeto"):
            nom_sujeto = _vision_campos["nom_sujeto"]
        if _vision_campos.get("id_sujeto"):
            id_sujeto = _vision_campos["id_sujeto"]
        if _vision_campos.get("base") and base == 0.0:
            base = float(_vision_campos["base"])
        if _vision_campos.get("ret") and ret == 0.0:
            ret = float(_vision_campos["ret"])

        if not _vision_campos and gemini_disponible():
            # Fallback textual solo cuando Vision no está disponible
            _nit_suj = id_sujeto if len(id_sujeto) == 14 else ""
            _dui_suj = id_sujeto if len(id_sujeto) == 9  else ""
            _campos_act = {
                "fecha"     : fecha,
                "nom_sujeto": nom_sujeto,
                "nit_sujeto": _nit_suj,
                "dui_sujeto": _dui_suj,
            }
            _corr_dict, gemini_correcciones = procesar_dte_con_gemini(
                texto_lineal,
                "sujetos_excluidos",
                _campos_act,
                {"nit": _nit_cliente_ctx, "nombre": _nom_cliente_ctx},
            )
            if _corr_dict.get("fecha"):
                fecha = _corr_dict["fecha"]
            if _corr_dict.get("nom_sujeto"):
                nom_sujeto = _corr_dict["nom_sujeto"]
            if _corr_dict.get("nit_sujeto"):
                id_sujeto = _corr_dict["nit_sujeto"]
            elif _corr_dict.get("dui_sujeto"):
                id_sujeto = _corr_dict["dui_sujeto"]

        return {
            "fecha"               : fecha,
            "id_sujeto"           : id_sujeto,
            "nom_sujeto"          : nom_sujeto,
            "tipo"                : tipo,
            "sello"               : sello,
            "gen"                 : gen,
            "num_control"         : num_control,
            "base"                : base,
            "ret"                 : ret,
            "gemini_correcciones" : gemini_correcciones,
            "_vision_campos"      : _vision_campos,
            "_vision_alertas"     : _vision_alertas,
            "_vision_audit"       : _vision_audit,
        }

    except pdfplumber.pdfminer.pdfparser.PDFSyntaxError:
        return {"error": "PDF inválido o corrupto."}
    except Exception as err:
        return {"error": str(err)}


# ─────────────────────────────────────────────
# GENERADORES DE EXCEL — 2 formatos
# ─────────────────────────────────────────────

def generar_excel_libro(df: pd.DataFrame, nombre_cliente: str) -> bytes:
    """
    LIBRO SUJETO EXCLUIDO — replica imagen 3:
    A(vacío) B(2) C(DUI/NIT) D(Nombre) E(Fecha) F(Sello) G(Núm.Control) H(Base) I(0.00) J...(1 1 4 5 5)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "LIBRO SUJETO EXCLUIDO"

    columnas = [
        ("",             "const_a",    4),
        ("",             "const_b",    5),
        ("DUI/NIT",      "id_sujeto", 14),
        ("Nombre",       "nom_sujeto",32),
        ("Fecha",        "fecha",     13),
        ("Sello Recep.", "sello",     44),   # ← SELLO en vez de código gen.
        ("Núm. Control", "num_control",38),
        ("Base ($)",     "base",      13),
        ("0.00",         "cero",       9),
        ("1","c1",4),("1","c2",4),("4","c3",4),("5","c4",4),("5","c5",4),
    ]

    header_fill = PatternFill("solid", fgColor="1A2C18")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    border_side = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side,  bottom=border_side)
    num_fmt  = '#,##0.00'
    alt_fill = PatternFill("solid", fgColor="F5F5F0")

    for col_idx, (header, _, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 16

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, (_, field, _) in enumerate(columnas, start=1):
            if field == "const_a":
                valor = ""
            elif field == "const_b":
                valor = 2
            elif field == "cero":
                valor = 0.00
            elif field.startswith("c") and field[1:].isdigit():
                valor = int(field[1:])
            else:
                valor = row.get(field, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical="center")
            if fill.fill_type:
                cell.fill = fill
            if field == "base":
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif field == "cero":
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif field == "fecha":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Totales
    total_row = len(df) + 2
    base_col  = [c[1] for c in columnas].index("base") + 1
    tot = ws.cell(row=total_row, column=base_col, value=df["base"].sum())
    tot.font = Font(bold=True)
    tot.number_format = num_fmt
    tot.alignment = Alignment(horizontal="right", vertical="center")
    tot.fill = PatternFill("solid", fgColor="A8E870")
    tot.border = cell_border

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_retencion(df: pd.DataFrame, nombre_cliente: str) -> bytes:
    """
    RETENCIÓN HONORARIOS — replica imagen 2:
    A(1) B(9300) C(Nombre) D(DUI/NIT) E(11) F(Base) G(Ret.) H...(ceros) ... Mes/Año
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RETENCIÓN HONORARIOS"

    columnas = [
        ("",         "c1",        4),
        ("Cód",      "c9300",     6),
        ("Nombre",   "nom_sujeto",32),
        ("DUI/NIT",  "id_sujeto", 13),
        ("",         "c11",        4),
        ("Base ($)", "base",      13),
        ("Ret. ($)", "ret",       12),
        ("0.00","cero1",8),("0.00","cero2",8),("0.00","cero3",8),
        ("0.00","cero4",8),("0.00","cero5",8),("0.00","cero6",8),("0.00","cero7",8),
        ("1","d1",4),("1","d2",4),("4","d3",4),("5","d4",4),
        ("Mes/Año","mesanio",10),
    ]

    header_fill = PatternFill("solid", fgColor="1A2C18")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    border_side = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side,  bottom=border_side)
    num_fmt  = '#,##0.00'
    alt_fill = PatternFill("solid", fgColor="F5F5F0")

    for col_idx, (header, _, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 16

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        fecha_str = row.get("fecha", "")
        mesanio   = ""
        mf = re.match(r"\d{2}\/(\d{2})\/(\d{4})", str(fecha_str))
        if mf:
            mesanio = f"{mf.group(1)}{mf.group(2)}"

        for col_idx, (_, field, _) in enumerate(columnas, start=1):
            if field == "c1":
                valor = 1
            elif field == "c9300":
                valor = 9300
            elif field == "c11":
                valor = 11
            elif field.startswith("cero"):
                valor = 0.00
            elif field.startswith("d") and field[1:].isdigit():
                valor = int(field[1:])
            elif field == "mesanio":
                valor = mesanio
            else:
                valor = row.get(field, "")

            cell  = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical="center")
            if fill.fill_type:
                cell.fill = fill
            if field in ("base", "ret") or field.startswith("cero"):
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif field in ("c1","c9300","c11","d1","d2","d3","d4"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Totales
    total_row = len(df) + 2
    base_col = [c[1] for c in columnas].index("base") + 1
    ret_col  = [c[1] for c in columnas].index("ret")  + 1
    for col_i, val in [(base_col, df["base"].sum()), (ret_col, df["ret"].sum())]:
        tot = ws.cell(row=total_row, column=col_i, value=val)
        tot.font = Font(bold=True)
        tot.number_format = num_fmt
        tot.alignment = Alignment(horizontal="right", vertical="center")
        tot.fill = PatternFill("solid", fgColor="A8E870")
        tot.border = cell_border

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# FORMATO HACIENDA — CASILLA 66 / ANEXO 5
# Compras a Sujetos Excluidos (13 columnas A-M)
# ══════════════════════════════════════════════════════════════

def construir_df_hacienda_c66(
    df: pd.DataFrame,
    tipo_op: str = "1",
    clasif: str  = "2",
    sector: str  = "4",
    tipo_cg: str = "1",
    periodo_feb2024: bool = True,
) -> pd.DataFrame:
    """
    Genera el DataFrame con la estructura exacta requerida por Hacienda
    para la Casilla 66 — Compras a Sujetos Excluidos (Anexo 5).

    Columnas A-M según Manual F-07:
      A  Tipo de Documento (1=NIT, 2=DUI, 3=Otro)
      B  Número NIT/DUI/Otro (sin guiones)
      C  Nombre, Razón Social o Denominación
      D  Fecha de Emisión DD/MM/AAAA
      E  Número de Serie (Sello de Recepción del DTE)
      F  Número de Documento (Código de Generación SIN guiones)
      G  Monto de la Operación
      H  Monto de la Retención IVA 13% (0.00 si no aplica)
      I  Tipo de Operación (desde Feb 2024; "0" si anterior)
      J  Clasificación (desde Feb 2024; "0" si anterior)
      K  Sector (desde Feb 2024; "0" si anterior)
      L  Tipo de Costo/Gasto (desde Feb 2024; "0" si anterior)
      M  Número de Anexo = 5
    """
    def _tipo_doc(id_str: str) -> str:
        clean = re.sub(r"[^0-9]", "", str(id_str))
        if len(clean) == 14:
            return "1"   # NIT
        if len(clean) == 9:
            return "2"   # DUI
        return "3"       # Otro

    def _iva13(base_val: float, t_op: str) -> str:
        # Solo aplica si la operación es Gravada (tipo 1) o Mixta (tipo 4)
        if t_op in ("1", "4"):
            return f"{round(float(base_val) * 0.13, 2):.2f}"
        return "0.00"

    out = pd.DataFrame()
    out["A"] = df["id_sujeto"].apply(_tipo_doc)
    out["B"] = df["id_sujeto"].astype(str).str.replace(r"[^0-9]", "", regex=True)
    out["C"] = df["nom_sujeto"].astype(str)
    out["D"] = df["fecha"].astype(str)
    out["E"] = df["sello"].astype(str)
    out["F"] = df["gen"].astype(str).str.replace("-", "", regex=False).str.upper()
    out["G"] = df["base"].map(lambda x: f"{float(x):.2f}")
    out["H"] = df["base"].apply(lambda x: _iva13(x, tipo_op))
    if periodo_feb2024:
        out["I"] = tipo_op
        out["J"] = clasif
        out["K"] = sector
        out["L"] = tipo_cg
    else:
        out["I"] = "0"
        out["J"] = "0"
        out["K"] = "0"
        out["L"] = "0"
    out["M"] = "5"
    return out


def to_csv_hacienda_c66(df_hacienda: pd.DataFrame) -> bytes:
    """CSV sin encabezados, listo para Hacienda (Casilla 66)."""
    return df_hacienda.to_csv(index=False, header=False).encode("utf-8")


_CODIGOS_INGRESO_F14 = {
    "0601 — Honorarios / Servicios Profesionales": "0601",
    "0701 — Agropecuario": "0701",
    "0801 — Arrendamiento Bienes Muebles": "0801",
    "0901 — Arrendamiento Bienes Inmuebles": "0901",
    "1001 — Otros Ingresos": "1001",
}

_TIPO_OP_F14 = {
    "1 — Gravada": "1",
    "2 — No Gravada": "2",
    "3 — Excluido / No Renta": "3",
    "4 — Mixta": "4",
}
_CLASIF_F14 = {"1 — Costo": "1", "2 — Gasto": "2"}
_SECTOR_F14 = {
    "1 — Industria": "1",
    "2 — Comercio": "2",
    "3 — Agropecuaria": "3",
    "4 — Servicios / Prof. / Artes y Oficios": "4",
}
_TIPO_CG_F14 = {
    "1 — Gastos de Venta sin Donación": "1",
    "2 — Gastos de Admón. sin Donación": "2",
    "3 — Gastos Financieros sin Donación": "3",
    "4 — Costo Art. Prod./Comprados Import./Internac.": "4",
    "5 — Costo Art. Prod./Comprados Interno": "5",
    "6 — Costos Indirectos de Fabricación": "6",
    "7 — Mano de Obra": "7",
}

_CORTE_DUI_F14 = pd.Timestamp(2022, 1, 1)


def construir_df_f14_isr(
    df: pd.DataFrame,
    codigo_ingreso: str = "0601",
    periodo: str = "012024",
    tipo_op: str = "3",
    clasif: str = "2",
    sector: str = "4",
    tipo_cg: str = "2",
    periodo_feb2024: bool = True,
) -> pd.DataFrame:
    """
    Genera el CSV Anexo F-14 (ISR Retenciones) — 23 columnas A-W.

    A  Domiciliado (1=Dom, 2=NoDom)
    B  Código de País (SLV para domiciliados)
    C  Apellidos, Nombres, Razón Social (max 100 chars)
    D  NIT/NIF (vacío si natural con DUI desde ene 2022)
    E  DUI (9 dígitos; vacío si jurídico o NIT sin DUI)
    F  Código de Ingreso (5 chars)
    G  Monto Devengado
    H  Monto Devengado por Bonificaciones/Gratificaciones (0)
    I  Impuesto Retenido
    J  Aguinaldo Exento (0)
    K  Aguinaldo Gravado (0)
    L  AFP (0)
    M  ISSS (0)
    N  INPEP (0)
    O  IPSFA (0)
    P  CEFAFA (0)
    Q  Bienestar Magisterial (0)
    R  ISS-VM (0)
    S  Tipo de Operación (desde Feb 2024; '0' si anterior)
    T  Clasificación (desde Feb 2024; '0' si anterior)
    U  Sector (desde Feb 2024; '0' si anterior)
    V  Tipo de Costo/Gasto (desde Feb 2024; '0' si anterior)
    W  Período MMYYYY
    """
    def _limpio(val: str) -> str:
        return re.sub(r"[^0-9]", "", str(val or ""))

    def _parse_fecha(fecha_str: str) -> pd.Timestamp:
        try:
            p = str(fecha_str).strip().split("/")
            if len(p) == 3:
                return pd.Timestamp(int(p[2]), int(p[1]), int(p[0]))
        except Exception:
            pass
        return pd.NaT

    col_d, col_e = [], []
    for id_raw, fecha_str in zip(df["id_sujeto"].astype(str), df["fecha"].astype(str)):
        limpio = _limpio(id_raw)
        ts = _parse_fecha(fecha_str)
        es_2022_plus = ts is not pd.NaT and ts >= _CORTE_DUI_F14
        if len(limpio) == 9 and es_2022_plus:
            col_d.append("")
            col_e.append(limpio)
        else:
            col_d.append(limpio)
            col_e.append("")

    out = pd.DataFrame()
    out["A"] = "1"
    out["B"] = "SLV"
    out["C"] = df["nom_sujeto"].astype(str).str[:100]
    out["D"] = col_d
    out["E"] = col_e
    out["F"] = codigo_ingreso
    out["G"] = df["base"].map(lambda x: f"{float(x):.2f}")
    out["H"] = "0.00"
    out["I"] = df["ret"].map(lambda x: f"{float(x):.2f}")
    for col_lbl in ("J", "K", "L", "M", "N", "O", "P", "Q", "R"):
        out[col_lbl] = "0.00"
    if periodo_feb2024:
        out["S"] = tipo_op
        out["T"] = clasif
        out["U"] = sector
        out["V"] = tipo_cg
    else:
        out["S"] = "0"
        out["T"] = "0"
        out["U"] = "0"
        out["V"] = "0"
    out["W"] = periodo
    return out


def to_csv_f14_isr(df_f14: pd.DataFrame) -> bytes:
    """CSV sin encabezados, listo para cargar en F-14 de Hacienda."""
    return df_f14.to_csv(index=False, header=False).encode("utf-8")


# ─────────────────────────────────────────────
# 5. ENCABEZADO
# ─────────────────────────────────────────────
col_logo, col_titulo = st.columns([1, 8])
with col_logo:
    st.markdown(
        "<h2 style='font-family: Courier New, monospace; color: #6AB040;"
        " letter-spacing: 3px; margin-top:8px;'>YN</h2>",
        unsafe_allow_html=True
    )
with col_titulo:
    st.title("⚖️ Extractor DTE — Sujetos Excluidos")

st.markdown(f"""
<div class="card-emisor">
    <strong>AGENTE DE RETENCIÓN:</strong> {cliente['nombre']}<br>
    <strong>NIT:</strong> {cliente['nit']}
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 6. SESSION STATE
# ─────────────────────────────────────────────
if 'db_sujetos'   not in st.session_state: st.session_state.db_sujetos   = pd.DataFrame()
if 'archivos_suj' not in st.session_state: st.session_state.archivos_suj = []

# ─────────────────────────────────────────────
# 7. SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Carga DTE-14")
    st.divider()

    archivos = st.file_uploader(
        "Arrastra DTE Sujetos Excluidos (PDF)",
        type="pdf",
        accept_multiple_files=True
    )

    procesar = st.button(
        "🚀 Procesar Documentos",
        type="primary",
        use_container_width=True,
        disabled=not archivos
    )
    limpiar = st.button(
        "🧹 Limpiar Memoria",
        type="secondary",
        use_container_width=True
    )

    if limpiar:
        st.session_state.db_sujetos   = pd.DataFrame()
        st.session_state.archivos_suj = []
        st.success("Memoria limpiada.")
        st.rerun()

    if procesar and archivos:
        procesados = set(st.session_state.archivos_suj)
        nuevos     = [f for f in archivos if f.name not in procesados]

        if not nuevos:
            st.info("ℹ️ Todos los archivos ya fueron procesados.")
        else:
            extracted  = []
            errores    = []
            bar        = st.progress(0)
            estado_txt = st.empty()
            total      = len(nuevos)

            # ── Pre-lectura en hilo principal ───────────────────────────────────
            nombres_y_bytes: list[tuple[str, bytes]] = [(f.name, f.read()) for f in nuevos]
            estado_txt.caption(
                f"⏳ Enviando {total} archivos en paralelo a Gemini..."
            )

            # ── Extracción paralela ─────────────────────────────────────────────
            fn_extraer = functools.partial(extraer_sujetos_nativo, cliente_activo=cliente)

            def _progreso_suj(comp: int, tot: int, fname: str) -> None:
                bar.progress(comp / tot)
                estado_txt.caption(f"⏳ {comp}/{tot} completados — `{fname}`")

            resultados = leer_y_procesar_lote(
                nombres_y_bytes, fn_extraer, progreso_cb=_progreso_suj,
            )

            # ── Clasificación secuencial ────────────────────────────────────────
            for fname, _fb, res in resultados:
                if "error" in res:
                    errores.append(f"❌ `{fname}` — {res['error']}")
                elif "error_tipo" in res:
                    errores.append(f"⚠️ `{fname}` — {res['error_tipo']}")
                else:
                    res["archivo"] = fname
                    _qa_alertas = validar_montos_sujetos(res)
                    _v_audit    = res.get("_vision_audit", {})
                    _confianza  = _v_audit.get("confianza", 100) if _v_audit else 100
                    res["_revision"] = "⚠️" if (_qa_alertas or _confianza < 65) else "✅"
                    extracted.append(res)

                st.session_state.archivos_suj.append(fname)

            estado_txt.empty()

            if extracted:
                new_df = pd.DataFrame(extracted)
                if st.session_state.db_sujetos.empty:
                    st.session_state.db_sujetos = new_df
                else:
                    st.session_state.db_sujetos = pd.concat(
                        [st.session_state.db_sujetos, new_df], ignore_index=True
                    )
                st.success(f"✅ {len(extracted)} documentos procesados.")

            if errores:
                st.warning(f"⚠️ {len(errores)} con error:")
                for e in errores:
                    st.markdown(e)

    if not st.session_state.db_sujetos.empty:
        st.divider()
        total_base = st.session_state.db_sujetos["base"].sum()
        total_ret  = st.session_state.db_sujetos["ret"].sum()
        st.markdown(f"**📄 Documentos:** `{len(st.session_state.db_sujetos)}`")
        st.markdown(f"**📊 Base Total:** `${total_base:,.2f}`")
        st.markdown(f"**⚖️ Ret. Total:** `${total_ret:,.2f}`")

# ─────────────────────────────────────────────
# 8. TABLA Y EXPORT
# ─────────────────────────────────────────────
if not st.session_state.db_sujetos.empty:
    df = st.session_state.db_sujetos.copy()

    st.markdown("#### 📋 Casilla 66 — Sujetos Excluidos (Base para F-14)")

    # Mostrar Sello en vez de Código de Generación
    COLS_DISPLAY = {
        "_revision"  : "QA",
        "fecha"      : "Fecha",
        "id_sujeto"  : "DUI/NIT",
        "nom_sujeto" : "Nombre",
        "tipo"       : "Tipo",
        "sello"      : "Sello de Recepción",
        "num_control": "Núm. Control",
        "base"       : "Base ($)",
        "ret"        : "Ret. Renta ($)",
    }
    cols_existentes = {k: v for k, v in COLS_DISPLAY.items() if k in df.columns}
    df_vista = df[list(cols_existentes.keys())].rename(columns=cols_existentes)

    st.dataframe(
        df_vista.style.format({
            "Base ($)"       : "${:,.2f}",
            "Ret. Renta ($)" : "${:,.2f}",
        }),
        hide_index=True,
        use_container_width=True,
        height=min(40 + len(df_vista) * 35, 600),
    )

    st.markdown(
        f"> **Base Total Operaciones:** `${df['base'].sum():,.2f}` &nbsp;|&nbsp;"
        f"**Retención Total 10%:** `${df['ret'].sum():,.2f}`"
    )

    st.markdown("---")

    # ── Configuración para Casilla 66 ────────────────────────────────────────
    st.markdown(
        "<span class='badge-sistema'>EXPORTAR — CASILLA 66</span>",
        unsafe_allow_html=True
    )
    st.markdown("")
    st.caption(
        "Configura los campos de clasificación requeridos por Hacienda "
        "(aplica desde Febrero 2024). El sistema los aplica igual a todos los registros. "
        "Puedes editar el CSV luego si hay documentos con clasificación distinta."
    )

    cc1, cc2, cc3, cc4, cc5 = st.columns(5)
    with cc1:
        sel_periodo = st.selectbox(
            "Periodo",
            ["Feb 2024 en adelante", "Anterior a Feb 2024"],
            key="c66_periodo",
        )
    with cc2:
        sel_tipo_op = st.selectbox(
            "Tipo de Operación (I)",
            ["1 — Gravada", "2 — No Gravada", "3 — Excluido/No Renta", "4 — Mixta"],
            key="c66_tipo_op",
        )
    with cc3:
        sel_clasif = st.selectbox(
            "Clasificación (J)",
            ["1 — Costo", "2 — Gasto"],
            key="c66_clasif",
        )
    with cc4:
        sel_sector = st.selectbox(
            "Sector (K)",
            ["1 — Industria", "2 — Comercio", "3 — Agropecuaria", "4 — Servicios/Prof."],
            key="c66_sector",
            index=3,
        )
    with cc5:
        sel_tipocg = st.selectbox(
            "Tipo Costo/Gasto (L)",
            [
                "1 — Gastos de Venta",
                "2 — Gastos de Admón.",
                "3 — Gastos Financieros",
                "4 — Costo Art. Import.",
                "5 — Costo Art. Interno",
                "6 — Costos Ind. Fab.",
                "7 — Mano de obra",
            ],
            key="c66_tipocg",
        )

    st.markdown("")

    col_h, col1, col2 = st.columns([2, 2, 2])

    with col_h:
        df_c66 = construir_df_hacienda_c66(
            df,
            tipo_op=sel_tipo_op[0],
            clasif=sel_clasif[0],
            sector=sel_sector[0],
            tipo_cg=sel_tipocg[0],
            periodo_feb2024=(sel_periodo == "Feb 2024 en adelante"),
        )
        st.markdown("**Formato Hacienda**")
        st.caption("CSV listo para cargar · Casilla 66 · Anexo 5")
        st.download_button(
            "📤 CSV Hacienda — Casilla 66",
            data=to_csv_hacienda_c66(df_c66),
            file_name=f"C66_SujetosExcluidos_{cliente['nombre'].replace(' ','_')}.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

    with col1:
        excel_libro = generar_excel_libro(df, cliente['nombre'])
        st.markdown("**Libro Sujeto Excluido**")
        st.caption("Excel formato auditoría F-14")
        st.download_button(
            "📥 Excel Libro F-14",
            data=excel_libro,
            file_name=f"Libro_Sujeto_Excluido_{cliente['nombre'].replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True
        )

    with col2:
        excel_ret = generar_excel_retencion(df, cliente['nombre'])
        st.markdown("**Retención Honorarios**")
        st.caption("Excel retención 10% ISR · F-14")
        st.download_button(
            "📥 Excel Ret. Honorarios",
            data=excel_ret,
            file_name=f"Retencion_Honorarios_{cliente['nombre'].replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True
        )

    # ── Exportar F-14 ISR (Hacienda CSV oficial) ─────────────────────────────
    st.markdown("---")
    st.markdown(
        "<span class='badge-sistema'>EXPORTAR — F-14 ISR (RETENCIONES RENTA)</span>",
        unsafe_allow_html=True
    )
    st.markdown("")
    st.caption(
        "Genera el archivo CSV en el formato exacto del Ministerio de Hacienda "
        "para cargar en la declaración F-14 (Retenciones de Impuesto sobre la Renta · Sujetos Domiciliados). "
        "23 columnas A-W. Columnas S-V aplican desde Febrero 2024."
    )

    f14_r1c1, f14_r1c2, f14_r1c3 = st.columns(3)
    with f14_r1c1:
        sel_cod_ingreso = st.selectbox(
            "F — Código de Ingreso",
            list(_CODIGOS_INGRESO_F14.keys()),
            key="f14_cod_ingreso",
        )
    with f14_r1c2:
        f14_mes = st.selectbox(
            "Período — Mes",
            ["01","02","03","04","05","06","07","08","09","10","11","12"],
            key="f14_mes",
        )
    with f14_r1c3:
        import datetime as _dt
        anio_actual = _dt.date.today().year
        f14_anio = st.selectbox(
            "Período — Año",
            [str(a) for a in range(anio_actual, 2020, -1)],
            key="f14_anio",
        )

    f14_periodo_str = f14_mes + f14_anio

    f14_c1, f14_c2, f14_c3, f14_c4, f14_c5 = st.columns(5)
    with f14_c1:
        f14_per_lbl = st.selectbox(
            "Período S-V",
            ["Feb 2024 en adelante", "Anterior a Feb 2024"],
            key="f14_periodo",
        )
    with f14_c2:
        f14_tipo_op_lbl = st.selectbox(
            "S — Tipo Operación",
            list(_TIPO_OP_F14.keys()),
            index=2,
            key="f14_tipo_op",
        )
    with f14_c3:
        f14_clasif_lbl = st.selectbox(
            "T — Clasificación",
            list(_CLASIF_F14.keys()),
            index=1,
            key="f14_clasif",
        )
    with f14_c4:
        f14_sector_lbl = st.selectbox(
            "U — Sector",
            list(_SECTOR_F14.keys()),
            index=3,
            key="f14_sector",
        )
    with f14_c5:
        f14_tipocg_lbl = st.selectbox(
            "V — Tipo Costo/Gasto",
            list(_TIPO_CG_F14.keys()),
            index=1,
            key="f14_tipocg",
        )

    df_f14 = construir_df_f14_isr(
        df,
        codigo_ingreso=_CODIGOS_INGRESO_F14[sel_cod_ingreso],
        periodo=f14_periodo_str,
        tipo_op=_TIPO_OP_F14[f14_tipo_op_lbl],
        clasif=_CLASIF_F14[f14_clasif_lbl],
        sector=_SECTOR_F14[f14_sector_lbl],
        tipo_cg=_TIPO_CG_F14[f14_tipocg_lbl],
        periodo_feb2024=(f14_per_lbl == "Feb 2024 en adelante"),
    )

    col_f14_dl, _ = st.columns([2, 4])
    with col_f14_dl:
        st.download_button(
            "📤 CSV Hacienda — F-14 ISR (Retenciones Renta)",
            data=to_csv_f14_isr(df_f14),
            file_name=f"F14_ISR_{cliente['nombre'].replace(' ','_')}_{f14_periodo_str}.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )
    st.caption(
        f"Período: **{f14_mes}/{f14_anio}** · Registros: **{len(df_f14)}** · "
        f"Total Devengado: **${df['base'].sum():,.2f}** · "
        f"Total ISR Retenido: **${df['ret'].sum():,.2f}**"
    )

else:
    st.markdown("""
    <div style="text-align:center; padding: 60px 20px; color: #4E7040;">
        <h3 style="color:#6AB040 !important;">📂 Sin documentos cargados</h3>
        <p style="color:#3A5830 !important;">Usa el panel lateral para cargar y procesar DTE-14 de sujetos excluidos.</p>
    </div>
    """, unsafe_allow_html=True)
