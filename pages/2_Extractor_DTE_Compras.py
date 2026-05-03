import streamlit as st
import pdfplumber
import pandas as pd
import re
import json
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# 1. PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title = "Extractor DTE Compras",
    layout     = "wide",
    page_icon  = "🛒"
)

# ─────────────────────────────────────────────
# 2. VERIFICACIÓN DE AUTENTICACIÓN
# ─────────────────────────────────────────────
if not st.session_state.get("autenticado"):
    st.warning("⚠️ Acceso denegado. Por favor, inicia sesión.")
    st.stop()

# ─────────────────────────────────────────────
# 3. ESTILOS — VERDE OLIVA
# ─────────────────────────────────────────────
ESTILO = """
<style>
  [data-testid="stAppViewContainer"],
  [data-testid="stHeader"]            { background-color: #0D0F07 !important; }
  [data-testid="stSidebar"]           { background-color: #141A08 !important;
                                        border-right: 1px solid #4A5520 !important; }
  h1,h2,h3,h4,h5,h6                  { color: #C8D87A !important; }
  p, label, span, li                  { color: #F0EDD8 !important; }

  div.stButton > button[kind="primary"] {
    background-color : #6B7A2A !important;
    border           : 1px solid #8A9A35 !important;
    border-radius    : 6px !important;
    transition       : 0.25s;
  }
  div.stButton > button[kind="primary"]:hover {
    background-color : #8A9A35 !important;
    transform        : scale(1.02);
  }
  div.stButton > button[kind="primary"] * {
    color: #FFFFFF !important; font-weight: bold !important;
  }
  div.stButton > button[kind="secondary"] {
    background-color : transparent !important;
    border           : 1px solid #4A5520 !important;
    border-radius    : 6px !important;
    transition       : 0.25s;
  }
  div.stButton > button[kind="secondary"]:hover { background-color: #1A2008 !important; }
  div.stButton > button[kind="secondary"] *     { color: #C8D87A !important; }

  div.stDownloadButton > button {
    background-color : #4A5520 !important;
    border           : 1px solid #6B7A2A !important;
    border-radius    : 6px !important;
    width            : 100%;
  }
  div.stDownloadButton > button * { color: #FFFFFF !important; font-weight: bold !important; }
  div.stDownloadButton > button:hover { background-color: #6B7A2A !important; }

  div[data-testid="stTextInput"] input,
  div[data-testid="stNumberInput"] input {
    background-color : #1A2008 !important;
    border           : 1px solid #4A5520 !important;
    border-radius    : 6px !important;
    color            : #F0EDD8 !important;
  }
  div[data-testid="stTextInput"] input:focus,
  div[data-testid="stNumberInput"] input:focus {
    border-color : #8A9A35 !important;
    box-shadow   : 0 0 0 2px rgba(138,154,53,0.25) !important;
  }
  div[data-testid="stSelectbox"] > div > div {
    background-color : #1A2008 !important;
    border           : 1px solid #4A5520 !important;
    border-radius    : 6px !important;
    color            : #F0EDD8 !important;
  }
  [data-testid="stDataFrame"]           { border: 1px solid #2A3010 !important; border-radius: 8px; }
  [data-testid="stDataFrame"] th        { background-color: #1A2008 !important; color: #C8D87A !important; }
  [data-testid="stDataFrame"] tr:hover  { background-color: #1A2008 !important; }

  button[data-baseweb="tab"]                       { color: #8A9A35 !important; }
  button[data-baseweb="tab"][aria-selected="true"] {
    border-bottom: 2px solid #8A9A35 !important;
    color: #F0EDD8 !important;
  }
  .stExpander { border: 1px solid #2A3010 !important; border-radius: 8px !important; }
  .stExpander header { color: #C8D87A !important; }

  hr { border-color: #4A5520 !important; opacity: 0.4; }

  .card-stat {
    background-color : #1A2008;
    border           : 1px solid #2A3010;
    border-radius    : 10px;
    padding          : 16px 20px;
    text-align       : center;
  }
  .card-stat .valor  { font-size: 1.8rem; font-weight: bold; color: #C8D87A !important; }
  .card-stat .etiq   { font-size: 0.8rem;  color: #6B7A2A !important; margin-top: 4px; }

  .alerta-revision {
    background-color : #2A1A08;
    border           : 1px solid #8A4A20;
    border-radius    : 8px;
    padding          : 10px 14px;
    margin-bottom    : 6px;
    font-size        : 0.88rem;
    color            : #F0CDA0 !important;
  }
  .alerta-revision strong { color: #FFB347 !important; }

  .badge-ok      { display:inline-block; padding:2px 8px; border-radius:10px;
                   background:#1A2A08; color:#8FCC30 !important;
                   border:1px solid #4A7A10; font-size:0.72rem; }
  .badge-rev     { display:inline-block; padding:2px 8px; border-radius:10px;
                   background:#2A1A08; color:#FFB347 !important;
                   border:1px solid #8A5020; font-size:0.72rem; }
  .badge-error   { display:inline-block; padding:2px 8px; border-radius:10px;
                   background:#2A0808; color:#FF7070 !important;
                   border:1px solid #8A2020; font-size:0.72rem; }
  .badge-calc    { display:inline-block; padding:2px 8px; border-radius:10px;
                   background:#1A1A2A; color:#A0AAFF !important;
                   border:1px solid #4A4A8A; font-size:0.72rem; }

  .footer-info {
    text-align  : center;
    font-size   : 0.75rem;
    color       : #4A5520 !important;
    margin-top  : 20px;
  }
</style>
"""
st.markdown(ESTILO, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 4. CONSTANTES
# ─────────────────────────────────────────────
TIPOS_VALIDOS_COMPRAS  = {"03", "05", "06"}
MAX_VALORES_LOOP       = 25
SUFIJOS_SOLOS          = {
    "S.A. DE C.V.", "C.V.", "SA DE CV", "LTDA", "LTDA.", "S.A.", "DE C.V."
}
PALABRAS_BASURA_NOMBRE = [
    "DOCUMENTO", "TRIBUTARIO", "ELECTRÓNICO", "REPRESENTACIÓN", "RECEPTOR",
    "CLIENTE", "EMISOR", "FACTURA", "CONSUMIDOR", "COMPROBANTE", "CÓDIGO",
    "SELLO", "VERSIÓN", "TRANSMISIÓN", "MINISTERIO", "HACIENDA", "COLONIA",
    "BOULEVARD", "CALLE", "AVENIDA", "MUNICIPIO", "ACTIVIDAD", "ECONOMICA",
    "ESTABLECIMIENTO", "SUCURSAL", "EFECTIVO", "TARJETA", "EMISIÓN",
    "GENERACIÓN", "TELÉFONO", "DIRECCIÓN", "GIRO", "FECHA", "HORA",
    "NÚMERO", "SERIE", "CONTROL", "VERSIÓN", "MODELO"
]
BASURA_ESTRICTA = ["@", ".COM", "WWW.", "HTTP", "CORREO", "EMAIL"]
INDICADORES_COMERCIAL = [
    "S.A.", "C.V.", "LTDA", "SOCIEDAD", "DISTRIBUIDORA", "FARMACIA",
    "GRUPO", "LABORATORIO", "INDUSTRIA", "CORPORACIÓN", "SERVICIOS",
    "COMERCIAL", "INVERSIONES", "CONSTRUCTORA", "IMPORTADORA",
    "EXPORTADORA", "TECNOLOGÍA", "SOLUCIONES", "EMPRESA", "COMPAÑIA"
]
ARCHIVO_PROVEEDORES = "data/proveedores.json"

# ─────────────────────────────────────────────
# 5. HELPERS DE PERSISTENCIA
# ─────────────────────────────────────────────
def cargar_proveedores_json() -> dict:
    """Carga la base de datos de proveedores desde disco."""
    if not os.path.exists(ARCHIVO_PROVEEDORES):
        return {}
    try:
        with open(ARCHIVO_PROVEEDORES, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError:
        st.warning("⚠️ El archivo de proveedores tiene un error de formato.")
        return {}
    except Exception as e:
        st.warning(f"⚠️ No se pudo cargar proveedores: {e}")
        return {}

def guardar_proveedores_json(db: dict) -> None:
    """Guarda la base de datos de proveedores en disco."""
    os.makedirs("data", exist_ok=True)
    with open(ARCHIVO_PROVEEDORES, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────
# 6. FUNCIONES DE EXTRACCIÓN
# ─────────────────────────────────────────────

def limpiar_monto(monto_str: str) -> float:
    """Convierte string de monto a float. Soporta formato anglosajón y europeo."""
    s = re.sub(r'[^\d.,]', '', str(monto_str).strip())
    if not s:
        return 0.0
    ultimo_coma  = s.rfind(',')
    ultimo_punto = s.rfind('.')
    try:
        if ultimo_coma > ultimo_punto:
            return float(s.replace('.', '').replace(',', '.'))
        elif ultimo_punto > ultimo_coma:
            return float(s.replace(',', ''))
        else:
            return float(re.sub(r'[^\d]', '', s))
    except ValueError:
        return 0.0


def extraer_y_formatear_fecha(texto: str) -> str:
    """Extrae y normaliza fecha de emisión con 4 estrategias. Retorna DD/MM/YYYY."""

    # Patrón 1 — Hacienda: YYYY-MM-DD
    m = re.search(
        r"\b(20[2-3]\d)\s*[-\/]\s*(0[1-9]|1[0-2])\s*[-\/]\s*(0[1-9]|[12]\d|3[01])\b",
        texto
    )
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"

    # Patrón 2 — DD/MM/YYYY o MM/DD/YYYY suelto
    m = re.search(
        r"\b(\d{1,2})\s*[\/\-\.]\s*(\d{1,2})\s*[\/\-\.]\s*(20[2-3]\d)\b",
        texto
    )
    if m:
        p1, p2, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if   p1 > 12 and p2 <= 12: return f"{p1:02d}/{p2:02d}/{y}"
        elif p2 > 12 and p1 <= 12: return f"{p2:02d}/{p1:02d}/{y}"
        elif p1 <= 12 and p2 <= 31: return f"{p1:02d}/{p2:02d}/{y}"

    # Patrón 3 — Etiqueta "Fecha de Emisión/Generación"
    m = re.search(
        r"(?:FECHA\s*(?:DE\s*)?(?:EMISI[OÓ]N|GENERACI[OÓ]N|EMISION|GENERACION))"
        r"[^\d]{0,25}(\d{1,2})[\/\-\.\s](\d{1,2})[\/\-\.\s](\d{4})",
        texto, re.I
    )
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if mo > 12 and d <= 12: d, mo = mo, d
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{d:02d}/{mo:02d}/{y}"

    # Patrón 4 — Fecha escrita: "13 de octubre de 2025"
    meses = {
        "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
        "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
    }
    m = re.search(
        r"(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
        r"septiembre|octubre|noviembre|diciembre)\s+(?:de\s+)?(\d{4})",
        texto, re.I
    )
    if m:
        d  = int(m.group(1))
        mo = meses.get(m.group(2).lower(), 0)
        y  = m.group(3)
        if mo and 1 <= d <= 31:
            return f"{d:02d}/{mo:02d}/{y}"

    return ""


def _normalizar_uuid(raw: str) -> str:
    """Normaliza UUID a formato estándar 8-4-4-4-12."""
    limpio = raw.replace("-", "").upper()
    if len(limpio) == 32:
        return f"{limpio[:8]}-{limpio[8:12]}-{limpio[12:16]}-{limpio[16:20]}-{limpio[20:]}"
    return ""


def _extraer_uuid(t_no_sp: str, t_clean: str) -> str:
    """Extrae UUID/Código de Generación con 4 estrategias en cascada."""

    # Estrategia 1 — URL de Hacienda: ?CODGEN=XXXX
    m = re.search(
        r"CODGEN=([A-F0-9]{8}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{12})",
        t_no_sp
    )
    if m: return _normalizar_uuid(m.group(1))

    # Estrategia 2 — Etiqueta explícita
    m = re.search(
        r"(?:C[OÓ]DIGO\s*DE\s*GENERACI[OÓ]N|COD\.?\s*GENERACI[OÓ]N)"
        r"[:\s]*([A-F0-9]{8}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{12})",
        t_no_sp, re.I
    )
    if m: return _normalizar_uuid(m.group(1))

    # Estrategia 3 — UUID flotante sin espacios
    m = re.search(
        r"([A-F0-9]{8}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{4}-?[A-F0-9]{12})",
        t_no_sp
    )
    if m: return _normalizar_uuid(m.group(1))

    # Estrategia 4 — UUID con espacios insertados por pdfplumber
    m = re.search(
        r"([A-F0-9]{8})\s*-?\s*([A-F0-9]{4})\s*-?\s*([A-F0-9]{4})"
        r"\s*-?\s*([A-F0-9]{4})\s*-?\s*([A-F0-9]{12})",
        t_clean, re.I
    )
    if m:
        return (
            f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            f"-{m.group(4)}-{m.group(5)}".upper()
        )

    return ""


def _extraer_nit_proveedor(
    texto_emisor: str,
    nit_receptor: str,
    dui_receptor: str
) -> list:
    """Extrae NITs/DUIs candidatos del emisor, excluyendo los del receptor."""
    patron = (
        r"\b\d{4}\s*-?\s*\d{6}\s*-?\s*\d{3}\s*-?\s*\d\b"
        r"|\b\d{14}\b"
        r"|\b\d{8}\s*-?\s*\d\b"
        r"|\b\d{9}\b"
    )
    encontrados = re.findall(patron, texto_emisor)
    limpios = list(dict.fromkeys(re.sub(r'[^0-9]', '', n) for n in encontrados))
    return [
        n for n in limpios
        if n not in (nit_receptor, dui_receptor) and len(n) >= 8
    ]


def _es_nombre_valido(nombre: str, palabras_cliente: list) -> bool:
    """Valida que un candidato sea un nombre de empresa aceptable."""
    if not nombre or len(nombre) < 4 or len(nombre) > 65:
        return False
    n_up = nombre.upper()
    if any(b in n_up for b in BASURA_ESTRICTA):
        return False
    if any(p in n_up for p in palabras_cliente):
        return False
    if n_up in SUFIJOS_SOLOS:
        return False
    if len(nombre) > 0 and sum(c.isdigit() for c in nombre) / len(nombre) > 0.4:
        return False
    return True


def _limpiar_nombre(nombre: str) -> str:
    """Elimina prefijos genéricos y caracteres basura de un nombre."""
    nombre = re.sub(
        r"^(?:RAZ[OÓ]N\s+SOCIAL|NOMBRE(?:\s+O\s+RAZ[OÓ]N\s+SOCIAL)?|"
        r"CLIENTE|NOMBRE\s+COMERCIAL|COMERCIAL|EMISOR)[\s:]*",
        "", nombre, flags=re.I
    ).strip()
    nombre = re.sub(r'^[\s\-_.,;:]+', '', nombre).strip()
    return nombre[:65]


def _extraer_nombre_proveedor(
    texto_emisor: str,
    nit_prov: str,
    proveedores_db: dict,
    cliente_activo: dict
) -> tuple:
    """
    Extrae nombre del proveedor con 5 estrategias en cascada.
    Retorna (nombre, es_nuevo).
    """
    # Paso 0 — Base de datos local
    if nit_prov and nit_prov in proveedores_db:
        return proveedores_db[nit_prov].get("nombre", ""), False

    palabras_cliente = cliente_activo.get('nombre', '').upper().split()[:3]

    # Estrategia 1 — Etiqueta "Nombre / Razón Social"
    m = re.search(
        r"(?:Nombre(?:\s+[Cc]omercial|\s+o\s+[Rr]az[oó]n\s+[Ss]ocial)?|"
        r"Raz[oó]n\s+Social)[:\s]+(.*?)"
        r"(?=\s*(?:NIT|NRC|Giro|Actividad|Direcci[oó]n|Tel[eé]fono|\n\n|$))",
        texto_emisor, re.I | re.DOTALL
    )
    if m:
        cand = re.sub(r'\s+', ' ', m.group(1)).strip()
        if _es_nombre_valido(cand, palabras_cliente):
            return _limpiar_nombre(cand), True

    # Estrategia 2 — Primera línea con indicador comercial
    lineas = texto_emisor.split('\n')
    for linea in lineas[:20]:
        L = linea.strip()
        if len(L) < 4: continue
        L_up = L.upper()
        if len(L) > 0 and sum(c.isdigit() for c in L) / len(L) > 0.35: continue
        if any(b in L_up for b in PALABRAS_BASURA_NOMBRE + BASURA_ESTRICTA): continue
        if any(p in L_up for p in palabras_cliente): continue
        if any(w in L_up for w in INDICADORES_COMERCIAL) and len(L) >= 5:
            nombre_cand = re.split(r'\s{3,}|(?:NIT|NRC)\s', L)[0].strip()
            if _es_nombre_valido(nombre_cand, palabras_cliente):
                return _limpiar_nombre(nombre_cand.upper()), True

    # Estrategia 3 — Línea después de "EMISOR"
    m_emisor = re.search(r"(?:EMISOR|PROVEEDOR)[:\s]*\n(.+)", texto_emisor, re.I)
    if m_emisor:
        cand = m_emisor.group(1).strip()
        if _es_nombre_valido(cand, palabras_cliente):
            return _limpiar_nombre(cand.upper()), True

    # Estrategia 4 — Bloque entre NRC y dirección
    m_bloque = re.search(
        r"NRC\s*:?\s*\d+[^\n]*\n(.+?)\n(?:Direcci[oó]n|Actividad|Giro)",
        texto_emisor, re.I | re.DOTALL
    )
    if m_bloque:
        cand = re.sub(r'\s+', ' ', m_bloque.group(1)).strip()
        if _es_nombre_valido(cand, palabras_cliente):
            return _limpiar_nombre(cand.upper()), True

    # Estrategia 5 — Cualquier línea limpia de 5-65 chars
    for linea in lineas[:35]:
        L = linea.strip().upper()
        if 5 <= len(L) <= 65:
            if len(L) > 0 and sum(c.isdigit() for c in L) / len(L) > 0.3: continue
            if any(b in L for b in PALABRAS_BASURA_NOMBRE + BASURA_ESTRICTA): continue
            if any(p in L for p in palabras_cliente): continue
            return _limpiar_nombre(L), True

    return "ESCRIBE EL NOMBRE AQUÍ", True


def _extraer_montos(t_clean: str, tipo: str) -> dict:
    """
    Extrae montos fiscales con 4 estrategias y fallback de cálculo.
    Retorna dict con exe, gra, iva, ret, perc, tot, iva_calc.
    """
    exe = gra = iva_val = ret = perc = tot = 0.0
    iva_calculado = False

    # FOVIAL + COTRANS (van a exentos)
    for termino in ["FOVIAL", "COTRANS"]:
        m_linea = re.search(rf"{termino}.{{0,50}}", t_clean, re.I)
        if m_linea:
            nums = re.findall(r"\d+[.,]\d{2}", m_linea.group(0))
            if nums:
                exe += max(limpiar_monto(n) for n in nums)
    exe = round(exe, 2)

    # Exentos declarados
    m_exe = re.search(
        r"(?:Ventas?\s+Exentas?|Total\s+Exento|Exento)[^\d]{0,30}"
        r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        t_clean, re.I
    )
    if m_exe:
        val_exe = limpiar_monto(m_exe.group(1))
        if val_exe > exe:
            exe = val_exe

    # Retención IVA
    m_ret = re.search(
        r"(?:IVA\s+)?(?:Retenido|Retenci[oó]n\s+IVA|Retenci[oó]n)[^\d]{0,20}"
        r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        t_clean, re.I
    )
    if m_ret:
        ret = limpiar_monto(m_ret.group(1))

    # IVA Percibido
    m_perc = re.search(
        r"(?:IVA\s+)?Percibido[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        t_clean, re.I
    )
    if m_perc:
        perc = limpiar_monto(m_perc.group(1))

    # Total — 7 patrones en cascada
    patrones_total = [
        r"TOTAL\s+A\s+PAGAR[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"TOTAL\s+PAGAR[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"MONTO\s+TOTAL[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"TOTAL\s+OPERACI[OÓ]N[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"VENTA\s+TOTAL[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"(?:\$|US\$)\s*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s*$",
        r"TOTAL[^\d]{0,10}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
    ]
    for patron in patrones_total:
        m_t = re.search(patron, t_clean, re.I | re.MULTILINE)
        if m_t:
            val = limpiar_monto(m_t.group(1))
            if val > 0:
                tot = val
                break

    # IVA — 5 patrones en cascada
    patrones_iva = [
        r"(?:D[EÉ]BITO|CR[EÉ]DITO)\s+FISCAL[^\d]{0,20}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"IVA\s+13%[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"13%\s+IVA[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"I\.?V\.?A\.?[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        r"Impuesto[^\d]{0,30}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
    ]
    for patron in patrones_iva:
        m_i = re.search(patron, t_clean, re.I)
        if m_i:
            val = limpiar_monto(m_i.group(1))
            if val > 0:
                iva_val = val
                break

    # Gravadas explícitas
    m_gra = re.search(
        r"(?:Ventas?\s+Gravadas?|Total\s+Grav[ao]do|Gravado)[^\d]{0,20}"
        r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        t_clean, re.I
    )
    if m_gra:
        gra = limpiar_monto(m_gra.group(1))

    # Triada matemática (límite de 25 valores para no hacer O(n³) sin control)
    if not (tot > 0 and iva_val > 0 and gra > 0):
        montos_raw = re.findall(
            r"(?:US\$?|\$)?\s*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})",
            t_clean
        )
        valores = sorted(
            {limpiar_monto(m) for m in montos_raw if limpiar_monto(m) > 0},
            reverse=True
        )[:MAX_VALORES_LOOP]

        encontrado = False
        for vt in valores:
            if encontrado: break
            for vg in valores:
                if vg >= vt: continue
                if encontrado: break
                for vi in valores:
                    if vi >= vg: continue
                    if (
                        abs(round(vg * 0.13, 2) - round(vi, 2)) <= 0.06
                        and abs(round(vg + vi + exe - ret, 2) - round(vt, 2)) <= 0.06
                    ):
                        gra, iva_val, tot = vg, vi, vt
                        encontrado = True
                        break

    # Fallback de cálculo cuando solo tenemos el total
    if tot > 0:
        if iva_val > 0 and gra == 0:
            gra = round(tot - iva_val - exe + ret, 2)
        elif iva_val == 0 and tipo == "03":
            gra           = round((tot + ret - exe) / 1.13, 2)
            iva_val       = round(tot + ret - exe - gra, 2)
            iva_calculado = True
        elif iva_val == 0 and tipo in ("05", "06"):
            base_neta = tot + ret - exe
            if base_neta > 0:
                gra           = round(base_neta / 1.13, 2)
                iva_val       = round(base_neta - gra, 2)
                iva_calculado = True

    return {
        "exe": exe,      "gra": max(gra, 0.0),
        "iva": iva_val,  "ret": ret,
        "perc": perc,    "tot": tot,
        "iva_calc": iva_calculado
    }


def _necesita_revision(res: dict) -> bool:
    """
    Umbral inteligente — solo manda a revisión si campos CRÍTICOS faltan.
    Reduce drásticamente el porcentaje de revisión manual.
    """
    sin_fecha  = not str(res.get('fecha', '')).strip()
    sin_gen    = not str(res.get('gen',   '')).strip()
    sin_total  = res.get('tot', 0.0) == 0.0
    sin_nombre = res.get('nom_prov', '').strip() in ('', 'ESCRIBE EL NOMBRE AQUÍ')

    if sin_total and sin_gen:   return True   # Sin datos fiscales útiles
    if sin_fecha:               return True   # Sin fecha no se puede declarar
    if sin_nombre and (sin_total or sin_gen): return True
    return False


def extraer_compras_nativo_pro(file_bytes: bytes, cliente_activo: dict) -> dict:
    """
    Extrae datos fiscales de un CCF/Nota de Crédito/Débito en PDF nativo.
    Versión mejorada — objetivo: reducir revisión manual de ~70% a <20%.
    """
    if not file_bytes or len(file_bytes) < 500:
        return {"error": "Archivo vacío o demasiado pequeño."}

    try:
        texto_lineal = ""
        texto_visual = ""
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return {"error": "PDF sin páginas."}
            for page in pdf.pages:
                texto_lineal += (page.extract_text(layout=False) or "") + "\n"
                texto_visual += (page.extract_text()              or "") + "\n"

        texto_completo = texto_lineal + "\n" + texto_visual

        if len(texto_completo.strip()) < 50:
            return {"error": "PDF de imagen — se requiere OCR."}

        # Normalizaciones
        t_clean = re.sub(r'[ \t]+', ' ', texto_completo)
        t_no_sp = re.sub(r'\s+',    '',  t_clean).upper()

        # ── Tipo DTE ─────────────────────────────────────────────────
        m_ctrl = re.search(r"(DTE-[0-9O]{2}-[A-Z0-9]+-[A-Z0-9]+)", t_no_sp)
        if not m_ctrl:
            return {"error_tipo": "No se encontró Número de Control DTE."}

        ctrl   = m_ctrl.group(1).replace("O", "0")
        m_tipo = re.search(r"DTE-(\d{2})", ctrl)
        tipo   = m_tipo.group(1) if m_tipo else "00"

        if tipo not in TIPOS_VALIDOS_COMPRAS:
            return {"error_tipo": f"DTE-{tipo} no admitido en Compras (válidos: 03, 05, 06)."}

        # ── UUID ─────────────────────────────────────────────────────
        gen = _extraer_uuid(t_no_sp, t_clean)

        # ── Fecha ─────────────────────────────────────────────────────
        fecha = extraer_y_formatear_fecha(t_clean)

        # ── Aislar sección EMISOR ─────────────────────────────────────
        partes = re.split(
            r"(?i)\b(?:RECEPTOR|CLIENTE\s*:|DATOS\s+DEL\s+RECEPTOR)\b",
            texto_lineal
        )
        texto_emisor = partes[0] if len(partes[0]) >= 80 else texto_lineal[:2000]

        # ── Identificadores del receptor ─────────────────────────────
        nit_receptor = re.sub(r'[^0-9]', '', cliente_activo.get('nit', ''))
        dui_receptor = re.sub(r'[^0-9]', '', cliente_activo.get('dui', ''))

        # ── NIT del proveedor ─────────────────────────────────────────
        candidatos       = _extraer_nit_proveedor(texto_emisor, nit_receptor, dui_receptor)
        proveedores_db   = cargar_proveedores_json()

        nit_prov = ""
        for n in candidatos:
            if n in proveedores_db:
                nit_prov = n
                break
        if not nit_prov and candidatos:
            nit_prov = candidatos[0]

        dui_prov = nit_prov if len(nit_prov) == 9 else ""

        # ── Nombre del proveedor ─────────────────────────────────────
        nom_prov, es_nuevo = _extraer_nombre_proveedor(
            texto_emisor, nit_prov, proveedores_db, cliente_activo
        )

        # ── Montos ───────────────────────────────────────────────────
        montos = _extraer_montos(t_clean, tipo)

        # ── Estado de coherencia matemática ──────────────────────────
        tot_calc = round(montos["gra"] + montos["iva"] + montos["exe"] - montos["ret"], 2)
        if   montos["tot"] == 0:
            estado = "Sin total"
        elif abs(tot_calc - montos["tot"]) > 0.15:
            estado = "Descuadre"
        else:
            estado = "OK"

        return {
            "fecha"    : fecha,
            "nit_prov" : nit_prov,
            "dui_prov" : dui_prov,
            "nom_prov" : nom_prov,
            "tipo"     : tipo,
            "ctrl"     : ctrl,
            "gen"      : gen,
            "exe"      : montos["exe"],
            "gra"      : montos["gra"],
            "iva"      : montos["iva"],
            "ret"      : montos["ret"],
            "perc"     : montos["perc"],
            "tot"      : montos["tot"],
            "estado"   : estado,
            "iva_calc" : montos["iva_calc"],
            "es_nuevo" : es_nuevo,
            "nit_nuevo": nit_prov if es_nuevo else "",
        }

    except Exception as err:
        return {"error": f"Error inesperado al procesar: {str(err)}"}


# ─────────────────────────────────────────────
# 7. GENERADOR DE EXCEL — HACIENDA F-07
# ─────────────────────────────────────────────
def generar_excel_hacienda(df: pd.DataFrame) -> bytes:
    """Genera el Excel con formato F-07 para Hacienda El Salvador."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = "F-07 Compras"

    # Paleta
    COLOR_HEADER    = "2A3010"
    COLOR_SUBHEADER = "1A2008"
    COLOR_FILA_PAR  = "141A08"
    COLOR_FILA_IMP  = "0D1005"
    COLOR_CALC      = "1A1A2A"
    FONT_HEADER     = Font(name="Calibri", bold=True,  color="C8D87A", size=10)
    FONT_BODY       = Font(name="Calibri", bold=False, color="F0EDD8", size=9)
    FONT_CALC       = Font(name="Calibri", bold=False, color="A0AAFF", size=9, italic=True)
    BORDER_THIN     = Border(
        left   = Side(style='thin', color='2A3010'),
        right  = Side(style='thin', color='2A3010'),
        top    = Side(style='thin', color='2A3010'),
        bottom = Side(style='thin', color='2A3010')
    )
    ALN_CTR = Alignment(horizontal='center', vertical='center')
    ALN_RGT = Alignment(horizontal='right',  vertical='center')
    ALN_LFT = Alignment(horizontal='left',   vertical='center')

    # Título
    ws.merge_cells('A1:O1')
    ws['A1'] = "ANEXO F-07 — LIBRO DE COMPRAS · LEARNIX DTE HUB"
    ws['A1'].font      = Font(name="Calibri", bold=True, color="C8D87A", size=12)
    ws['A1'].fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    ws['A1'].alignment = ALN_CTR
    ws.row_dimensions[1].height = 24

    # Encabezados
    encabezados = [
        "N°", "FECHA", "NIT PROVEEDOR", "DUI PROVEEDOR",
        "NOMBRE PROVEEDOR", "TIPO DTE", "CTRL DTE",
        "CÓDIGO GENERACIÓN", "EXENTAS", "GRAVADAS",
        "IVA CRÉDITO", "IVA CALC?", "RETENCIÓN",
        "IVA PERCIBIDO", "TOTAL"
    ]
    ws.append(encabezados)
    fila_h = ws.max_row
    for col_idx, _ in enumerate(encabezados, start=1):
        cell = ws.cell(row=fila_h, column=col_idx)
        cell.font      = FONT_HEADER
        cell.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = ALN_CTR
        cell.border    = BORDER_THIN
    ws.row_dimensions[fila_h].height = 18

    # Datos
    for idx, row in enumerate(df.itertuples(index=False), start=1):
        fila_vals = [
            idx,
            getattr(row, 'fecha',    ''),
            getattr(row, 'nit_prov', ''),
            getattr(row, 'dui_prov', ''),
            getattr(row, 'nom_prov', ''),
            f"DTE-{getattr(row, 'tipo', '')}",
            getattr(row, 'ctrl',     ''),
            getattr(row, 'gen',      ''),
            getattr(row, 'exe',  0.0),
            getattr(row, 'gra',  0.0),
            getattr(row, 'iva',  0.0),
            "SÍ (calc.)" if getattr(row, 'iva_calc', False) else "",
            getattr(row, 'ret',  0.0),
            getattr(row, 'perc', 0.0),
            getattr(row, 'tot',  0.0),
        ]
        ws.append(fila_vals)
        r = ws.max_row
        ws.row_dimensions[r].height = 15

        bg = COLOR_FILA_PAR if idx % 2 == 0 else COLOR_FILA_IMP
        for col_idx, val in enumerate(fila_vals, start=1):
            cell = ws.cell(row=r, column=col_idx)
            # Color especial si IVA fue calculado
            if col_idx == 11 and getattr(row, 'iva_calc', False):
                cell.fill = PatternFill("solid", fgColor=COLOR_CALC)
                cell.font = FONT_CALC
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = FONT_BODY

            cell.border = BORDER_THIN
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment     = ALN_RGT
            elif col_idx in (1, 6, 12):
                cell.alignment = ALN_CTR
            else:
                cell.alignment = ALN_LFT

    # Fila de totales
    r_tot = ws.max_row + 1
    ws.cell(row=r_tot, column=4, value="TOTALES").font = Font(
        name="Calibri", bold=True, color="C8D87A", size=10
    )
    ws.cell(row=r_tot, column=4).fill = PatternFill("solid", fgColor=COLOR_HEADER)

    for c_idx, col_name in [(9,"exe"),(10,"gra"),(11,"iva"),(13,"ret"),(14,"perc"),(15,"tot")]:
        val  = df[col_name].sum() if col_name in df.columns else 0.0
        cell = ws.cell(row=r_tot, column=c_idx, value=round(val, 2))
        cell.font          = Font(name="Calibri", bold=True, color="C8D87A", size=10)
        cell.fill          = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.number_format = '#,##0.00'
        cell.alignment     = ALN_RGT
        cell.border        = BORDER_THIN

    # Anchos de columna
    anchos = [4,12,16,14,40,9,30,38,12,12,12,10,12,14,13]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# 8. INICIALIZACIÓN DE SESSION STATE
# ─────────────────────────────────────────────
if "db_compras"       not in st.session_state: st.session_state["db_compras"]       = pd.DataFrame()
if "revision_compras" not in st.session_state: st.session_state["revision_compras"] = []
if "proveedores_nuevos" not in st.session_state: st.session_state["proveedores_nuevos"] = {}

# ─────────────────────────────────────────────
# 9. ENCABEZADO
# ─────────────────────────────────────────────
col_ico, col_ttl = st.columns([1, 10])
with col_ico:
    st.markdown(
        "<h1 style='font-size:2.5rem; margin:0; padding-top:4px;'>🛒</h1>",
        unsafe_allow_html=True
    )
with col_ttl:
    st.markdown(
        "<h2 style='margin:0; padding-top:8px;'>Extractor DTE Compras</h2>"
        "<p style='color:#6B7A2A; font-size:0.88rem; margin:0;'>"
        "Procesa CCF · Notas de Crédito · Notas de Débito &nbsp;|&nbsp; "
        "Genera Anexo F-07</p>",
        unsafe_allow_html=True
    )

st.divider()

# ─────────────────────────────────────────────
# 10. VERIFICAR CLIENTE ACTIVO
# ─────────────────────────────────────────────
cliente_activo = st.session_state.get("cliente_activo")
if not cliente_activo:
    st.warning("⚠️ **Sin empresa activa.** Ve al Dashboard y selecciona una empresa antes de procesar.")
    st.stop()

# Card del cliente activo
st.markdown(
    f"<div style='background:#1A2008; border-left:4px solid #8A9A35; "
    f"border-radius:8px; padding:10px 16px; margin-bottom:16px;'>"
    f"<span style='color:#6B7A2A; font-size:0.75rem; letter-spacing:1px;'>PROCESANDO PARA</span><br>"
    f"<strong style='color:#C8D87A; font-size:1rem;'>{cliente_activo.get('nombre','—')}</strong>"
    f"<span style='color:#6B7A2A; font-size:0.85rem; margin-left:12px;'>"
    f"NIT: {cliente_activo.get('nit','—')}</span></div>",
    unsafe_allow_html=True
)

# ─────────────────────────────────────────────
# 11. CARGA DE ARCHIVOS
# ─────────────────────────────────────────────
archivos_pdf = st.file_uploader(
    "Carga tus PDFs de compras (CCF, Notas de Crédito/Débito)",
    type        = ["pdf"],
    accept_multiple_files = True,
    help        = "Formatos admitidos: DTE-03 (CCF), DTE-05 (Nota de Crédito), DTE-06 (Nota de Débito)"
)

col_proc, col_lim = st.columns([3, 1])
with col_proc:
    procesar = st.button(
        "⚙️ Procesar PDFs",
        type             = "primary",
        use_container_width = True,
        disabled         = (not archivos_pdf)
    )
with col_lim:
    if st.button("🗑️ Limpiar Todo", use_container_width=True):
        st.session_state["db_compras"]        = pd.DataFrame()
        st.session_state["revision_compras"]  = []
        st.session_state["proveedores_nuevos"] = {}
        st.rerun()

# ─────────────────────────────────────────────
# 12. PROCESAMIENTO
# ─────────────────────────────────────────────
if procesar and archivos_pdf:
    progreso   = st.progress(0, text="Iniciando procesamiento...")
    resultados = []
    revision   = []
    prov_nuevos = {}

    for i, archivo in enumerate(archivos_pdf):
        nombre_archivo = archivo.name
        progreso.progress(
            (i + 1) / len(archivos_pdf),
            text=f"Procesando {i+1}/{len(archivos_pdf)}: {nombre_archivo}"
        )

        file_bytes = archivo.read()
        res        = extraer_compras_nativo_pro(file_bytes, cliente_activo)

        # Errores de extracción (PDF de imagen u otros)
        if "error" in res:
            revision.append({
                "archivo" : nombre_archivo,
                "motivo"  : f"Error técnico: {res['error']}",
                "datos"   : {}
            })
            continue

        # DTE de tipo no admitido
        if "error_tipo" in res:
            revision.append({
                "archivo" : nombre_archivo,
                "motivo"  : res["error_tipo"],
                "datos"   : {}
            })
            continue

        # Umbral inteligente de revisión
        if _necesita_revision(res):
            motivos = []
            if not res.get("fecha"):             motivos.append("Sin fecha")
            if not res.get("gen"):               motivos.append("Sin UUID")
            if res.get("tot", 0.0) == 0.0:      motivos.append("Sin total")
            if res.get("nom_prov", "").strip() in ("", "ESCRIBE EL NOMBRE AQUÍ"):
                motivos.append("Sin nombre proveedor")
            revision.append({
                "archivo" : nombre_archivo,
                "motivo"  : " | ".join(motivos) if motivos else "Revisión requerida",
                "datos"   : res
            })
            continue

        # Guardar proveedor nuevo detectado
        if res.get("es_nuevo") and res.get("nit_nuevo"):
            prov_nuevos[res["nit_nuevo"]] = {
                "nombre"  : res["nom_prov"],
                "nit"     : res["nit_nuevo"],
                "archivo" : nombre_archivo
            }

        res["archivo"] = nombre_archivo
        resultados.append(res)

    progreso.empty()

    # Acumular resultados
    if resultados:
        df_nuevo = pd.DataFrame(resultados)
        if st.session_state["db_compras"].empty:
            st.session_state["db_compras"] = df_nuevo
        else:
            st.session_state["db_compras"] = pd.concat(
                [st.session_state["db_compras"], df_nuevo],
                ignore_index=True
            ).drop_duplicates(subset=["gen"], keep="last")

    st.session_state["revision_compras"]   = revision
    st.session_state["proveedores_nuevos"] = prov_nuevos

    # Resumen flash
    total_pdfs = len(archivos_pdf)
    ok_count   = len(resultados)
    rev_count  = len(revision)

    c_ok, c_rev, c_tot = st.columns(3)
    with c_ok:
        st.markdown(f"""
        <div class="card-stat">
            <div class="valor" style="color:#8FCC30 !important;">{ok_count}</div>
            <div class="etiq">✅ Extraídos automáticamente</div>
        </div>""", unsafe_allow_html=True)
    with c_rev:
        st.markdown(f"""
        <div class="card-stat">
            <div class="valor" style="color:#FFB347 !important;">{rev_count}</div>
            <div class="etiq">⚠️ Requieren revisión</div>
        </div>""", unsafe_allow_html=True)
    with c_tot:
        pct = round(ok_count / total_pdfs * 100) if total_pdfs else 0
        st.markdown(f"""
        <div class="card-stat">
            <div class="valor">{pct}%</div>
            <div class="etiq">🎯 Tasa de extracción automática</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("")

    if prov_nuevos:
        st.info(
            f"🆕 **{len(prov_nuevos)} proveedor(es) nuevo(s)** detectado(s). "
            f"Revísalos en la pestaña **Proveedores Nuevos**."
        )

# ─────────────────────────────────────────────
# 13. PESTAÑAS DE RESULTADOS
# ─────────────────────────────────────────────
db      = st.session_state["db_compras"]
rev     = st.session_state["revision_compras"]
p_nuevos = st.session_state["proveedores_nuevos"]

tab_bd, tab_rev, tab_prov, tab_sum = st.tabs([
    f"📋 Base de Datos ({len(db)})",
    f"⚠️ Revisión Manual ({len(rev)})",
    f"🆕 Proveedores Nuevos ({len(p_nuevos)})",
    "📊 Resumen Financiero"
])

# ── TAB 1: Base de Datos ──────────────────────────────────────────────
with tab_bd:
    if db.empty:
        st.info("📂 Aún no hay documentos procesados. Carga PDFs y presiona **Procesar PDFs**.")
    else:
        st.markdown(f"**{len(db)} documento(s)** procesado(s) correctamente.")
        st.markdown("")

        # Editor con columnas clave
        cols_mostrar = [
            c for c in
            ["archivo","fecha","nom_prov","nit_prov","tipo","gra","iva","exe","ret","perc","tot","estado","iva_calc","gen"]
            if c in db.columns
        ]
        df_edit = st.data_editor(
            db[cols_mostrar].copy(),
            use_container_width = True,
            num_rows            = "dynamic",
            column_config       = {
                "archivo"  : st.column_config.TextColumn("📄 Archivo",      width=220),
                "fecha"    : st.column_config.TextColumn("📅 Fecha",        width=100),
                "nom_prov" : st.column_config.TextColumn("🏢 Proveedor",    width=220),
                "nit_prov" : st.column_config.TextColumn("🪪 NIT",          width=140),
                "tipo"     : st.column_config.TextColumn("📝 Tipo",         width=70),
                "gra"      : st.column_config.NumberColumn("💰 Gravadas",   width=100, format="$%.2f"),
                "iva"      : st.column_config.NumberColumn("🏦 IVA",        width=90,  format="$%.2f"),
                "exe"      : st.column_config.NumberColumn("🔓 Exentas",    width=90,  format="$%.2f"),
                "ret"      : st.column_config.NumberColumn("✂️ Retención",  width=90,  format="$%.2f"),
                "perc"     : st.column_config.NumberColumn("📌 Percibido",  width=90,  format="$%.2f"),
                "tot"      : st.column_config.NumberColumn("💵 Total",      width=100, format="$%.2f"),
                "estado"   : st.column_config.TextColumn("🔍 Estado",       width=90),
                "iva_calc" : st.column_config.CheckboxColumn("IVA calc.",   width=70),
                "gen"      : st.column_config.TextColumn("🔑 UUID",         width=290),
            },
            hide_index = True,
            key        = "editor_compras"
        )

        # Guardar ediciones
        if st.button("💾 Guardar Ediciones", type="primary"):
            edited_df = pd.DataFrame(df_edit)
            for c in db.columns:
                if c not in cols_mostrar:
                    edited_df[c] = db[c].values[:len(edited_df)] if len(db) >= len(edited_df) else None
            st.session_state["db_compras"] = edited_df
            st.success("✅ Cambios guardados correctamente.")
            st.rerun()

        st.divider()

        # Exportar
        col_xl, col_csv = st.columns(2)
        with col_xl:
            xl_bytes = generar_excel_hacienda(db)
            nombre_empresa = re.sub(
                r'[^A-Za-z0-9_]', '_',
                cliente_activo.get('nombre', 'empresa')
            )[:20]
            st.download_button(
                label            = "📥 Descargar Excel F-07 (Hacienda)",
                data             = xl_bytes,
                file_name        = f"F07_Compras_{nombre_empresa}.xlsx",
                mime             = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width = True,
                type             = "primary"
            )
        with col_csv:
            csv_bytes = db.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
            st.download_button(
                label            = "📄 Descargar CSV",
                data             = csv_bytes,
                file_name        = f"Compras_{nombre_empresa}.csv",
                mime             = "text/csv",
                use_container_width = True
            )

# ── TAB 2: Revisión Manual ────────────────────────────────────────────
with tab_rev:
    if not rev:
        st.success("🎉 Sin documentos pendientes de revisión manual.")
    else:
        st.warning(
            f"⚠️ **{len(rev)} documento(s)** requieren revisión manual. "
            f"Completa los campos faltantes y muévelos a la Base de Datos."
        )
        st.markdown("")

        for i, item in enumerate(rev):
            motivo   = item.get("motivo", "Sin descripción")
            archivo  = item.get("archivo", f"documento_{i+1}")
            datos    = item.get("datos", {})

            with st.expander(f"📄 {archivo}  —  ⚠️ {motivo}", expanded=False):
                col_f, col_t = st.columns([1, 1])
                with col_f:
                    fecha_v = st.text_input(
                        "Fecha (DD/MM/YYYY)", value=datos.get("fecha", ""),
                        key=f"rev_fecha_{i}"
                    )
                    nom_v = st.text_input(
                        "Nombre del Proveedor", value=datos.get("nom_prov", ""),
                        key=f"rev_nom_{i}"
                    )
                    nit_v = st.text_input(
                        "NIT Proveedor", value=datos.get("nit_prov", ""),
                        key=f"rev_nit_{i}"
                    )
                with col_t:
                    gra_v  = st.number_input("Ventas Gravadas",   value=float(datos.get("gra",  0.0)), step=0.01, key=f"rev_gra_{i}")
                    iva_v  = st.number_input("IVA Crédito Fiscal", value=float(datos.get("iva",  0.0)), step=0.01, key=f"rev_iva_{i}")
                    exe_v  = st.number_input("Ventas Exentas",     value=float(datos.get("exe",  0.0)), step=0.01, key=f"rev_exe_{i}")
                    tot_v  = st.number_input("Total a Pagar",      value=float(datos.get("tot",  0.0)), step=0.01, key=f"rev_tot_{i}")
                    ret_v  = st.number_input("Retención IVA",      value=float(datos.get("ret",  0.0)), step=0.01, key=f"rev_ret_{i}")
                    perc_v = st.number_input("IVA Percibido",      value=float(datos.get("perc", 0.0)), step=0.01, key=f"rev_perc_{i}")

                gen_v  = st.text_input(
                    "Código de Generación (UUID)", value=datos.get("gen", ""),
                    key=f"rev_gen_{i}"
                )
                tipo_v = st.selectbox(
                    "Tipo DTE",
                    options=["03 — CCF", "05 — Nota de Crédito", "06 — Nota de Débito"],
                    index  = {"03":0,"05":1,"06":2}.get(datos.get("tipo","03"), 0),
                    key    = f"rev_tipo_{i}"
                )

                if st.button(
                    "✅ Mover a Base de Datos",
                    type="primary", key=f"rev_mover_{i}",
                    use_container_width=True
                ):
                    nuevo_reg = {
                        "archivo"  : archivo,
                        "fecha"    : fecha_v,
                        "nom_prov" : nom_v,
                        "nit_prov" : nit_v,
                        "dui_prov" : datos.get("dui_prov", ""),
                        "tipo"     : tipo_v.split(" ")[0],
                        "ctrl"     : datos.get("ctrl", ""),
                        "gen"      : gen_v,
                        "exe"      : exe_v,
                        "gra"      : gra_v,
                        "iva"      : iva_v,
                        "iva_calc" : datos.get("iva_calc", False),
                        "ret"      : ret_v,
                        "perc"     : perc_v,
                        "tot"      : tot_v,
                        "estado"   : "Manual",
                    }
                    df_nuevo = pd.DataFrame([nuevo_reg])
                    if st.session_state["db_compras"].empty:
                        st.session_state["db_compras"] = df_nuevo
                    else:
                        st.session_state["db_compras"] = pd.concat(
                            [st.session_state["db_compras"], df_nuevo],
                            ignore_index=True
                        )
                    st.session_state["revision_compras"].pop(i)
                    st.success(f"✅ '{archivo}' movido a la base de datos.")
                    st.rerun()

# ── TAB 3: Proveedores Nuevos ─────────────────────────────────────────
with tab_prov:
    if not p_nuevos:
        st.info("📭 No hay proveedores nuevos detectados en esta sesión.")
    else:
        st.markdown(
            f"**{len(p_nuevos)} proveedor(es) nuevo(s)** detectado(s). "
            f"Guárdalos en el Directorio para que se reconozcan automáticamente en el futuro."
        )
        st.markdown("")

        db_prov_actual = cargar_proveedores_json()

        for nit_k, datos_p in p_nuevos.items():
            with st.expander(f"🏢 {datos_p.get('nombre', '?')}  —  NIT: {nit_k}", expanded=True):
                col_n, col_nit = st.columns(2)
                with col_n:
                    nom_conf = st.text_input(
                        "Confirmar Nombre",
                        value=datos_p.get("nombre", ""),
                        key=f"prov_nom_{nit_k}"
                    )
                with col_nit:
                    nit_conf = st.text_input(
                        "Confirmar NIT",
                        value=nit_k,
                        key=f"prov_nit_{nit_k}"
                    )

                if st.button(
                    f"💾 Guardar en Directorio",
                    key=f"prov_guardar_{nit_k}",
                    type="primary",
                    use_container_width=True
                ):
                    db_prov_actual[nit_conf] = {
                        "nombre" : nom_conf,
                        "nit"    : nit_conf
                    }
                    guardar_proveedores_json(db_prov_actual)
                    del st.session_state["proveedores_nuevos"][nit_k]
                    st.success(f"✅ Proveedor '{nom_conf}' guardado en el Directorio.")
                    st.rerun()

# ── TAB 4: Resumen Financiero ─────────────────────────────────────────
with tab_sum:
    if db.empty:
        st.info("📊 Procesa documentos para ver el resumen financiero.")
    else:
        st.markdown("#### 💰 Totales del Período")
        st.markdown("")

        total_gra  = db["gra"].sum()  if "gra"  in db.columns else 0.0
        total_iva  = db["iva"].sum()  if "iva"  in db.columns else 0.0
        total_exe  = db["exe"].sum()  if "exe"  in db.columns else 0.0
        total_ret  = db["ret"].sum()  if "ret"  in db.columns else 0.0
        total_perc = db["perc"].sum() if "perc" in db.columns else 0.0
        total_tot  = db["tot"].sum()  if "tot"  in db.columns else 0.0
        n_docs     = len(db)
        n_calc     = db["iva_calc"].sum() if "iva_calc" in db.columns else 0

        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">${total_gra:,.2f}</div>
                <div class="etiq">💰 Total Compras Gravadas</div>
            </div>""", unsafe_allow_html=True)
        with r1c2:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">${total_iva:,.2f}</div>
                <div class="etiq">🏦 Total IVA Crédito Fiscal</div>
            </div>""", unsafe_allow_html=True)
        with r1c3:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">${total_exe:,.2f}</div>
                <div class="etiq">🔓 Total Compras Exentas</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("")

        r2c1, r2c2, r2c3 = st.columns(3)
        with r2c1:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor" style="color:#FFB347 !important;">${total_ret:,.2f}</div>
                <div class="etiq">✂️ Total Retenciones IVA</div>
            </div>""", unsafe_allow_html=True)
        with r2c2:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor" style="color:#A0AAFF !important;">${total_perc:,.2f}</div>
                <div class="etiq">📌 Total IVA Percibido</div>
            </div>""", unsafe_allow_html=True)
        with r2c3:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">${total_tot:,.2f}</div>
                <div class="etiq">💵 Total General Pagado</div>
            </div>""", unsafe_allow_html=True)

        st.divider()
        st.markdown("#### 📈 Métricas de Calidad")
        st.markdown("")

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">{n_docs}</div>
                <div class="etiq">📄 Documentos procesados</div>
            </div>""", unsafe_allow_html=True)
        with mc2:
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor">{int(n_calc)}</div>
                <div class="etiq">🧮 IVA calculado (estimado)</div>
            </div>""", unsafe_allow_html=True)
        with mc3:
            pct_auto = round(
                n_docs / (n_docs + len(rev)) * 100
            ) if (n_docs + len(rev)) > 0 else 0
            color_pct = "#8FCC30" if pct_auto >= 80 else "#FFB347" if pct_auto >= 50 else "#FF7070"
            st.markdown(f"""
            <div class="card-stat">
                <div class="valor" style="color:{color_pct} !important;">{pct_auto}%</div>
                <div class="etiq">🎯 Tasa de extracción automática</div>
            </div>""", unsafe_allow_html=True)

        # Tabla por tipo DTE
        if "tipo" in db.columns:
            st.divider()
            st.markdown("#### 📂 Desglose por Tipo DTE")
            resumen_tipo = db.groupby("tipo").agg(
                Documentos = ("tot", "count"),
                Gravadas   = ("gra", "sum"),
                IVA        = ("iva", "sum"),
                Exentas    = ("exe", "sum"),
                Total      = ("tot", "sum")
            ).reset_index()
            resumen_tipo.rename(columns={"tipo": "Tipo DTE"}, inplace=True)
            st.dataframe(
                resumen_tipo,
                use_container_width = True,
                hide_index          = True,
                column_config       = {
                    "Tipo DTE"  : st.column_config.TextColumn(width=100),
                    "Documentos": st.column_config.NumberColumn(width=100),
                    "Gravadas"  : st.column_config.NumberColumn(format="$%.2f", width=130),
                    "IVA"       : st.column_config.NumberColumn(format="$%.2f", width=120),
                    "Exentas"   : st.column_config.NumberColumn(format="$%.2f", width=120),
                    "Total"     : st.column_config.NumberColumn(format="$%.2f", width=130),
                }
            )

# ─────────────────────────────────────────────
# 14. FOOTER
# ─────────────────────────────────────────────
st.divider()
st.markdown(
    "<p class='footer-info'>Learnix DTE Hub v2.0 &nbsp;·&nbsp; "
    "Extractor DTE Compras &nbsp;·&nbsp; El Salvador<br>"
    "Todos los documentos se procesan localmente. "
    "Ningún dato es enviado a servidores externos.</p>",
    unsafe_allow_html=True
)
