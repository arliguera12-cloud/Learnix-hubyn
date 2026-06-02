import functools
import streamlit as st
import pdfplumber
import pandas as pd
import re
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from styles import DARK_PRO_CSS
from utils.concurrent_processor import leer_y_procesar_lote
from utils.pdf_utils import (
    safe_str,
    normalizar_unicode,
    limpiar_monto,
    extraer_y_formatear_fecha,
    extraer_texto_pdf,
)
from utils.ai_utils import (
    gemini_disponible,
    gemini_ultimo_error,
    procesar_dte_con_gemini,
)
from utils.gemini_vision import (
    extraer_dte_con_vision,
    vision_disponible,
    vision_ultimo_error,
)
from utils.qa_utils import (
    mostrar_banner_qa,
    mostrar_indicador_vision,
    validar_montos_retenciones,
)
from utils.qr_reader import extraer_datos_qr as _extraer_qr

# ─────────────────────────────────────────────
# 1. PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(page_title="Extractor DTE · Retenciones", layout="wide", page_icon="✂️")

# ─────────────────────────────────────────────
# 2. ESTILOS
# ─────────────────────────────────────────────
st.markdown(DARK_PRO_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 3. SEGURIDAD — Multi-tenant SaaS
# ─────────────────────────────────────────────
from utils.auth_guard import check_auth
check_auth()

if not st.session_state.get("cliente_activo"):
    st.warning("⚠️ Debes seleccionar un Cliente Activo en el Dashboard antes de extraer Retenciones.")
    st.stop()

cliente = st.session_state.cliente_activo

# ─────────────────────────────────────────────
# 4. FUNCIONES AUXILIARES
# ─────────────────────────────────────────────
def cargar_proveedores_json() -> dict:
    """
    Retorna dict combinado {nit: {nombre, nrc}}:
    - Catálogo global (proveedores_globales) como base
    - Catálogo privado de la org encima con prioridad
    """
    try:
        from utils.local_db import cargar_proveedores_combinados
        return cargar_proveedores_combinados()
    except Exception:
        return {}








def extraer_sello(texto_original: str) -> str:
    """Extrae el Sello de Recepción MH: alfanumérico ~40 chars sin guiones."""

    def _valido(v: str) -> bool:
        v = v.strip().upper()
        return (
            25 <= len(v) <= 60
            and "-" not in v
            and re.search(r"[A-Z]", v)
            and re.search(r"[0-9]", v)
            and not v.startswith("DTE")
        )

    t_ns = re.sub(r"\s+", "", texto_original).upper()

    # 1. Etiqueta en la misma línea (valor inmediato)
    m = re.search(
        r"(?:Sello\s+de\s+Recepci[oó]n|SelloRecibido|Sello\s+Recibido)"
        r"[\s:=\-]*([A-Z0-9]{25,60})",
        texto_original, re.I,
    )
    if m and _valido(m.group(1)):
        return m.group(1).strip().upper()

    # 2. Sin espacios (PDF que colapsa whitespace entre etiqueta y valor)
    for pat_ns in (
        r"SELLODERECEPCI[O0]N[:\-=]?([A-Z0-9]{25,60})",
        r"SELLORECIBIDO[:\-=]?([A-Z0-9]{25,60})",
        r"SELLORECEIBIDO[:\-=]?([A-Z0-9]{25,60})",
        r"SELLORECE[PC]CION[:\-=]?([A-Z0-9]{25,60})",
    ):
        m = re.search(pat_ns, t_ns)
        if m and _valido(m.group(1)):
            return m.group(1).upper()

    # 3. JSON-like embebido en el texto del PDF
    for pat_json in (
        r'"[Ss]ello[Rr]ecibido"\s*:\s*"([A-Z0-9]{25,60})"',
        r"'[Ss]ello[Rr]ecibido'\s*:\s*'([A-Z0-9]{25,60})'",
        r"[Ss]ello[Rr]ecibido\s*[=:]\s*\"?([A-Z0-9]{25,60})\"?",
        r"respuesta[Hh]acienda[^\"']{0,120}[Ss]ello[Rr]ecibido[\"'\s:=]+([A-Z0-9]{25,60})",
        r"response[Mm][Hh][^\"']{0,120}[Ss]ello[Rr]ecibido[\"'\s:=]+([A-Z0-9]{25,60})",
    ):
        m = re.search(pat_json, texto_original)
        if m and _valido(m.group(1)):
            return m.group(1).upper()

    # 4. Etiqueta en una línea, sello en la siguiente (layout vertical)
    lineas = texto_original.splitlines()
    for i, linea in enumerate(lineas):
        if re.search(
            r"[Ss]ello\s+(?:de\s+)?[Rr]ecepci[oó]n|[Ss]ello\s*[Rr]ecibido",
            linea,
        ):
            for sig in lineas[i + 1 : i + 5]:
                cand = re.sub(r"[^A-Z0-9]", "", sig.strip().upper())
                if _valido(cand):
                    return cand

    # 5. Cerca de "Fecha Procesado", "Fecha y Hora de Generación" (zona del sello MH)
    for pat_ctx in (
        r"(?:Fecha\s+[Yy]\s+Hora\s+de\s+Generaci[oó]n|Fecha\s+Procesad[oa]|"
        r"Procesado\s+(?:por\s+)?MH|FechaHora\s*Recepci[oó]n)"
        r"[^\n]{0,200}?([A-Z0-9]{30,50})",
    ):
        m = re.search(pat_ctx, texto_original, re.I | re.S)
        if m and _valido(m.group(1)):
            return m.group(1).upper()

    # 6. Standalone 36-44 chars alfanumérico (no es UUID puro de 32 hex)
    for cand in re.findall(r"(?<![A-Z0-9])([A-Z0-9]{36,44})(?![A-Z0-9])", t_ns):
        if _valido(cand) and len(cand) != 32:   # 32 = UUID sin guiones (solo hex)
            return cand

    # 7. Heurística de inicio por año (sello suele comenzar con el año de emisión)
    for linea in lineas:
        mc = re.match(r"^\s*(20[2-9]\d[A-Z0-9]{28,38})\s*$", linea, re.I)
        if mc and _valido(mc.group(1)):
            return mc.group(1).upper()

    return ""


def extraer_retencion_nativa(file_bytes: bytes, cliente_activo: dict) -> dict:
    if not file_bytes or len(file_bytes) < 512:
        return {"error": "Archivo vacío o corrupto."}

    # ── Vision-First: extraer con IA antes de pdfplumber ─────────────────────
    _nit_cliente_ctx = re.sub(r'[^0-9]', '', cliente_activo.get('nit', ''))
    _nom_cliente_ctx = cliente_activo.get('nombre', '')

    gemini_correcciones: list[str] = []
    _vision_campos: dict  = {}
    _vision_alertas: list = []
    _vision_audit: dict   = {}

    if vision_disponible():
        _vision_campos, _vision_alertas, _vision_audit = extraer_dte_con_vision(
            file_bytes,
            "retenciones",
            {"nit": _nit_cliente_ctx, "nombre": _nom_cliente_ctx},
        )
        gemini_correcciones = [
            f"Vision: {a}" for a in _vision_alertas
        ] if _vision_alertas else (
            [f"Vision extrajo {len(_vision_campos)} campo(s)"]
            if _vision_campos else []
        )

    try:
        texto_lineal, texto_visual = extraer_texto_pdf(file_bytes)
        texto_completo = texto_lineal + "\n" + texto_visual

        if len(texto_completo.strip()) < 50 and not _vision_campos.get("num_control"):
            return {"error": "PDF de imagen — sin texto extraíble."}

        t_clean = re.sub(r'[ \t]+', ' ', texto_completo)
        t_no_sp = re.sub(r'\s+', '', t_clean).upper()

        m_ctrl = re.search(r"(DTE-[0-9O]{2}-[A-Z0-9]{1,20}-\d{9,18})", t_no_sp)
        tipo   = "07"
        if m_ctrl:
            ctrl   = m_ctrl.group(1).replace("O", "0")
            m_tipo = re.search(r"DTE-(\d{2})", ctrl)
            if m_tipo:
                tipo = m_tipo.group(1)

        if tipo != "07":
            return {"error_tipo": f"Documento DTE-{tipo}. Solo se admiten DTE-07 (Retenciones)."}

        nit_cliente = re.sub(r'[^0-9]', '', cliente_activo.get('nit', ''))

        gen = ""
        m_uuid = re.search(
            r"([A-Fa-f0-9]{8}-?[A-Fa-f0-9]{4}-?[A-Fa-f0-9]{4}-?[A-Fa-f0-9]{4}-?[A-Fa-f0-9]{12})",
            t_no_sp
        )
        if m_uuid:
            raw = m_uuid.group(1).replace("-", "")
            if len(raw) == 32:
                gen = f"{raw[:8]}-{raw[8:12]}-{raw[12:16]}-{raw[16:20]}-{raw[20:]}".upper()

        sello = extraer_sello(t_clean)
        fecha = extraer_y_formatear_fecha(t_clean)

        nit_prov = ""
        patron_ids = (
            r"\b\d{4}\s*-?\s*\d{6}\s*-?\s*\d{3}\s*-?\s*\d\b"
            r"|\b\d{14}\b"
        )
        # Buscar en texto_completo (con espacios, captura NITs con separadores)
        ids_raw  = re.findall(patron_ids, texto_completo)
        # Fallback: buscar 14 dígitos consecutivos en texto sin espacios
        ids_raw += re.findall(r'\d{14}', t_no_sp)
        ids_limp   = list(dict.fromkeys(re.sub(r'[^0-9]', '', n) for n in ids_raw))
        candidatos = [n for n in ids_limp if n != nit_cliente and len(n) == 14]

        proveedores_db = cargar_proveedores_json()
        for n in candidatos:
            if n in proveedores_db:
                nit_prov = n
                break

        if not nit_prov and candidatos:
            nit_prov = candidatos[0]

        base, ret = 0.0, 0.0

        m_base = re.search(
            r"(?:Monto\s+[Ss]ujeto|[Ss]ujeto\s+a\s+[Rr]etenci[oó]n|"
            r"[Tt]otal\s+[Mm]onto\s+[Ss]ujeto(?:\s+a\s+[Rr]etener?)?)"
            r"[^\d$]{0,30}\$?\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?|\d+)",
            t_clean, re.I
        )
        m_ret = re.search(
            r"(?:[Tt]otal\s+IVA\s+[Rr]etenido|[Tt]otal\s+IVA\s+[Rr]eteni"
            r"|[Ii]mpuesto\s+[Rr]etenido|[Rr]etenci[oó]n\s+1%|[Mm]onto\s+[Rr]etenci[oó]n)"
            r"[^\d$]{0,30}\$?\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?|\d+)",
            t_clean, re.I
        )

        if m_base:
            base = limpiar_monto(m_base.group(1))
        if m_ret:
            ret = limpiar_monto(m_ret.group(1))

        if base == 0.0:
            montos_raw = re.findall(
                r"(?:US\$?|\$)\s*(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?|\d{2,}(?:[.,]\d{1,2})?)",
                t_clean
            )
            valores = sorted(
                list({limpiar_monto(m) for m in montos_raw if limpiar_monto(m) > 0}),
                reverse=True
            )
            for v in valores:
                ret_calc = round(v * 0.01, 2)
                if any(abs(r - ret_calc) <= 0.05 for r in valores if r < v):
                    base = v
                    ret  = ret_calc
                    break

        if base > 0 and ret == 0:
            ret = round(base * 0.01, 2)

        if ret > 0 and base == 0:
            base = round(ret * 100, 2)

        # ── Aplicar Vision con prioridad sobre regex ──────────────────────────
        if _vision_campos.get("fecha"):
            fecha    = _vision_campos["fecha"]
        if _vision_campos.get("nit_prov"):
            nit_prov = _vision_campos["nit_prov"]
        if _vision_campos.get("base") and base == 0.0:
            base = float(_vision_campos["base"])
        if _vision_campos.get("ret") and ret == 0.0:
            ret = float(_vision_campos["ret"])

        # Sello: aplicar Vision si regex no lo encontró o encontró uno corto
        v_sello = str(_vision_campos.get("sello_recepcion") or "").strip()
        if len(v_sello) >= 25 and "-" not in v_sello and len(sello) < 25:
            sello = v_sello

        if not _vision_campos and gemini_disponible():
            # Fallback textual solo cuando Vision no está disponible
            _campos_act = {"fecha": fecha, "nit_prov": nit_prov}
            _corr_dict, gemini_correcciones = procesar_dte_con_gemini(
                texto_lineal,
                "retenciones",
                _campos_act,
                {"nit": _nit_cliente_ctx, "nombre": _nom_cliente_ctx},
            )
            if _corr_dict.get("fecha"):
                fecha    = _corr_dict["fecha"]
            if _corr_dict.get("nit_prov"):
                nit_prov = _corr_dict["nit_prov"]

        # ── QR ES EL REY: sobreescribe gen si el QR encontró datos ─────────────
        try:
            _qr = _extraer_qr(file_bytes)
            if _qr.get("codigo_generacion"):
                gen = _qr["codigo_generacion"].upper()
        except Exception:
            pass

        return {
            "nit_prov"            : nit_prov,
            "fecha"               : fecha,
            "tipo"                : tipo,
            "sello"               : sello,
            "gen"                 : gen,
            "base"                : base,
            "ret"                 : ret,
            "estado"              : 7,
            "gemini_correcciones" : gemini_correcciones,
            "_vision_campos"      : _vision_campos,
            "_vision_alertas"     : _vision_alertas,
            "_vision_audit"       : _vision_audit,
        }

    except pdfplumber.pdfminer.pdfparser.PDFSyntaxError:
        return {"error": "PDF inválido o corrupto."}
    except Exception as err:
        return {"error": str(err)}


# ══════════════════════════════════════════════════════════════
# FORMATO HACIENDA — CASILLA 162 / ANEXO 7
# Retención IVA 1% efectuada al declarante (9 columnas A-I)
# ══════════════════════════════════════════════════════════════

def construir_df_hacienda_c162(df: pd.DataFrame) -> pd.DataFrame:
    """
    Genera el DataFrame con la estructura exacta requerida por el portal
    de Hacienda El Salvador para la Casilla 162 (Retención IVA 1%).

    Columnas A-I según Manual F-07:
      A  NIT del Agente (14 dig, sin guiones; vacío si se usa DUI en H)
      B  Fecha de Emisión DD/MM/AAAA
      C  Tipo de Documento (07 = Comprobante de Retención)
      D  Serie (Sello de Recepción del DTE)
      E  Número de Documento (Código de Generación SIN guiones)
      F  Monto Sujeto
      G  Monto Retención 1%
      H  DUI del Agente (9 dig; vacío si se usa NIT en A)
      I  Número de Anexo = 7
    """
    out = pd.DataFrame()
    out["A"] = df["nit_prov"].astype(str).str.replace(r"[^0-9]", "", regex=True)
    out["B"] = df["fecha"].astype(str)
    out["C"] = df["tipo"].astype(str)
    out["D"] = df["sello"].astype(str)
    # E: UUID sin guiones
    out["E"] = df["gen"].astype(str).str.replace("-", "", regex=False).str.upper()
    out["F"] = df["base"].map(lambda x: f"{float(x):.2f}")
    out["G"] = df["ret"].map(lambda x: f"{float(x):.2f}")
    out["H"] = ""   # DUI del agente (persona natural); vacío cuando se usa NIT
    out["I"] = "7"
    return out


def to_csv_hacienda(df_hacienda: pd.DataFrame) -> bytes:
    """CSV sin encabezados, sin separador de miles, listo para Hacienda."""
    return df_hacienda.to_csv(index=False, header=False).encode("utf-8")


def generar_excel(df: pd.DataFrame, nombre_cliente: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Retenciones F-14"

    columnas = [
        ("NIT Proveedor",      "nit_prov", 18),
        ("Fecha",              "fecha",    14),
        ("Tipo",               "tipo",      6),
        ("Sello de Recepción", "sello",    44),
        ("Código de Gen.",     "gen",      38),
        ("Base ($)",           "base",     14),
        ("Retención ($)",      "ret",      14),
        ("Estado",             "estado",    8),
    ]

    header_fill = PatternFill("solid", fgColor="1A2C18")
    header_font = Font(bold=True, color="A8E870", size=10)
    border_side  = Side(style="thin", color="2E4828")
    cell_border  = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )
    num_fmt  = '#,##0.00'
    alt_fill = PatternFill("solid", fgColor="F2F7EF")

    for col_idx, (header, _, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, (_, field, _) in enumerate(columnas, start=1):
            valor = row.get(field, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical="center")
            if fill.fill_type:
                cell.fill = fill
            if field in ("base", "ret"):
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif field == "estado":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field in ("tipo", "fecha"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = len(df) + 2
    ws.cell(row=total_row, column=5, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")

    base_col = [c[1] for c in columnas].index("base") + 1
    ret_col  = [c[1] for c in columnas].index("ret")  + 1

    tot_base = ws.cell(row=total_row, column=base_col, value=df["base"].sum())
    tot_ret  = ws.cell(row=total_row, column=ret_col,  value=df["ret"].sum())
    for cell in (tot_base, tot_ret):
        cell.font          = Font(bold=True, color="111E12")
        cell.number_format = num_fmt
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.fill          = PatternFill("solid", fgColor="A8E870")
        cell.border        = cell_border

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# 5. ENCABEZADO
# ─────────────────────────────────────────────
col_logo, col_titulo = st.columns([1, 8])
with col_logo:
    st.markdown(
        "<h2 style='font-family: Courier New, monospace; color: #A8E870;"
        " letter-spacing: 3px; margin-top:8px;'>YN</h2>",
        unsafe_allow_html=True
    )
with col_titulo:
    st.title("✂️ Extractor DTE — Retenciones 1%")

st.markdown(f"""
<div class="card-emisor">
    <div class="label">Agente de Retención</div>
    <div class="nombre">{cliente['nombre']}</div>
    <div class="nit">NIT: {cliente['nit']}</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 6. SESSION STATE
# ─────────────────────────────────────────────
if 'db_ret'       not in st.session_state: st.session_state.db_ret       = pd.DataFrame()
if 'archivos_ret' not in st.session_state: st.session_state.archivos_ret = []

# ─────────────────────────────────────────────
# 7. SIDEBAR — CARGA Y PROCESAMIENTO
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Carga DTE-07")
    st.caption("Comprobantes de Retención 1% — Formulario F-14")
    st.divider()

    archivos = st.file_uploader(
        "Arrastra los PDFs aquí",
        type="pdf",
        accept_multiple_files=True
    )

    procesar = st.button(
        "🚀 Procesar Retenciones",
        type="primary",
        use_container_width=True,
        disabled=not archivos
    )
    limpiar = st.button(
        "🧹 Limpiar Memoria",
        type="secondary",
        use_container_width=True
    )

    if limpiar:
        st.session_state.db_ret       = pd.DataFrame()
        st.session_state.archivos_ret = []
        st.success("Memoria limpiada.")
        st.rerun()

    if procesar and archivos:
        procesados = set(st.session_state.archivos_ret)
        nuevos     = [f for f in archivos if f.name not in procesados]

        if not nuevos:
            st.info("ℹ️ Todos los archivos ya fueron procesados.")
        else:
            extracted  = []
            errores    = []
            bar        = st.progress(0)
            estado_txt = st.empty()
            total      = len(nuevos)

            # ── Pre-lectura en hilo principal ───────────────────────────────────
            nombres_y_bytes: list[tuple[str, bytes]] = [(f.name, f.read()) for f in nuevos]
            estado_txt.caption(
                f"⏳ Enviando {total} archivos en paralelo a Gemini..."
            )

            # ── Extracción paralela ─────────────────────────────────────────────
            fn_extraer = functools.partial(extraer_retencion_nativa, cliente_activo=cliente)

            def _progreso_ret(comp: int, tot: int, fname: str) -> None:
                bar.progress(comp / tot)
                estado_txt.caption(f"⏳ {comp}/{tot} completados — `{fname}`")

            resultados = leer_y_procesar_lote(
                nombres_y_bytes, fn_extraer, progreso_cb=_progreso_ret,
            )

            # ── Clasificación secuencial ────────────────────────────────────────
            for fname, _fb, res in resultados:
                if "error" in res:
                    errores.append(f"❌ `{fname}` — {res['error']}")
                elif "error_tipo" in res:
                    errores.append(f"⚠️ `{fname}` — {res['error_tipo']}")
                else:
                    res["archivo"] = fname
                    _qa_alertas = validar_montos_retenciones(res)
                    _v_audit    = res.get("_vision_audit", {})
                    _confianza  = _v_audit.get("confianza", 100) if _v_audit else 100
                    res["_revision"] = "⚠️" if (_qa_alertas or _confianza < 65) else "✅"
                    extracted.append(res)

                st.session_state.archivos_ret.append(fname)

            estado_txt.empty()

            if extracted:
                new_df = pd.DataFrame(extracted)
                if st.session_state.db_ret.empty:
                    st.session_state.db_ret = new_df
                else:
                    st.session_state.db_ret = pd.concat(
                        [st.session_state.db_ret, new_df], ignore_index=True
                    )
                st.success(f"✅ {len(extracted)} retenciones procesadas.")

            if errores:
                st.warning(f"⚠️ {len(errores)} con error:")
                for e in errores:
                    st.markdown(e)

    if not st.session_state.db_ret.empty:
        st.divider()
        total_base = st.session_state.db_ret["base"].sum()
        total_ret  = st.session_state.db_ret["ret"].sum()
        st.markdown(f"**📄 Documentos:** `{len(st.session_state.db_ret)}`")
        st.markdown(f"**📊 Base Total:** `${total_base:,.2f}`")
        st.markdown(f"**✂️ Ret. Total:** `${total_ret:,.2f}`")

# ─────────────────────────────────────────────
# 8. TABLA, FILTROS Y EXPORT
# ─────────────────────────────────────────────
if not st.session_state.db_ret.empty:
    df = st.session_state.db_ret.copy()

    # ── Panel de Filtros ────────────────────────────────────────────────────
    st.markdown('<div class="filter-panel">', unsafe_allow_html=True)
    st.markdown('<span class="filter-title">🔍 Filtros — Libro de Retenciones F-14</span>', unsafe_allow_html=True)

    ff1, ff2, ff3, ff4, ff5 = st.columns([2, 1, 1, 1, 1])
    with ff1:
        busqueda_ret = st.text_input(
            "busqueda_ret", label_visibility="collapsed",
            placeholder="Buscar NIT proveedor o Sello/UUID…"
        )
    with ff2:
        fd_desde = st.date_input("Desde", value=None, format="DD/MM/YYYY", key="ret_fd")
    with ff3:
        fd_hasta = st.date_input("Hasta", value=None, format="DD/MM/YYYY", key="ret_fh")
    with ff4:
        base_min = st.number_input("Base mín. ($)", min_value=0.0, value=0.0, step=10.0, key="ret_bm")
    with ff5:
        base_max = st.number_input("Base máx. ($)", min_value=0.0, value=0.0, step=100.0,
                                    key="ret_bx", help="0 = sin límite")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Aplicar filtros ──────────────────────────────────────────────────────
    df_fil = df.copy()

    if busqueda_ret:
        t = busqueda_ret.strip()
        mask = (
            df_fil['nit_prov'].str.contains(t, case=False, na=False) |
            df_fil['sello'].str.contains(t, case=False, na=False)    |
            df_fil['gen'].str.contains(t, case=False, na=False)
        )
        df_fil = df_fil[mask]

    def _dmy_ts(fecha_str: str):
        try:
            p = str(fecha_str).strip().split('/')
            if len(p) == 3:
                return pd.Timestamp(int(p[2]), int(p[1]), int(p[0]))
        except Exception:
            pass
        return pd.NaT

    if (fd_desde or fd_hasta) and 'fecha' in df_fil.columns:
        df_fil['_fts'] = df_fil['fecha'].apply(_dmy_ts)
        if fd_desde:
            df_fil = df_fil[df_fil['_fts'] >= pd.Timestamp(fd_desde)]
        if fd_hasta:
            df_fil = df_fil[df_fil['_fts'] <= pd.Timestamp(fd_hasta)]
        df_fil = df_fil.drop(columns=['_fts'], errors='ignore')

    if base_min > 0:
        df_fil = df_fil[df_fil['base'] >= base_min]
    if base_max > 0:
        df_fil = df_fil[df_fil['base'] <= base_max]

    # Badge de resultados
    n_tot = len(df)
    n_fil = len(df_fil)
    filtros_activos = sum([
        bool(busqueda_ret), bool(fd_desde), bool(fd_hasta),
        bool(base_min > 0), bool(base_max > 0),
    ])
    badge_extra = (f'<span class="active-filters"> · {filtros_activos} filtro(s) activo(s)</span>'
                   if filtros_activos else "")
    st.markdown(
        f'<div class="results-badge"><span class="cnt">{n_fil}</span> de {n_tot} registros{badge_extra}</div>',
        unsafe_allow_html=True
    )

    # ── Tabla principal ──────────────────────────────────────────────────────
    st.markdown("#### 📋 Libro de Retenciones — Base para Formulario F-14")

    COLS_DISPLAY = {
        "_revision": "QA",
        "nit_prov" : "NIT Proveedor",
        "fecha"    : "Fecha",
        "tipo"     : "Tipo",
        "sello"    : "Sello de Recepción",
        "gen"      : "Código de Generación",
        "base"     : "Base ($)",
        "ret"      : "Retención ($)",
        "estado"   : "Estado",
    }

    cols_existentes = {k: v for k, v in COLS_DISPLAY.items() if k in df_fil.columns}
    df_vista = df_fil[list(cols_existentes.keys())].rename(columns=cols_existentes)

    st.dataframe(
        df_vista.style.format({
            "Base ($)"      : "${:,.2f}",
            "Retención ($)" : "${:,.2f}",
        }),
        hide_index=True,
        use_container_width=True,
        height=min(40 + len(df_vista) * 35, 600),
    )

    # Resumen de totales
    base_tot = df_fil['base'].sum() if 'base' in df_fil.columns else 0.0
    ret_tot  = df_fil['ret'].sum()  if 'ret'  in df_fil.columns else 0.0
    st.markdown(
        f"> **Base Total:** `${base_tot:,.2f}` &nbsp;|&nbsp;"
        f"**Retención Total 1%:** `${ret_tot:,.2f}`"
    )

    st.markdown("---")

    # ── Descarga ─────────────────────────────────────────────────────────────
    st.markdown(
        "<span class='badge-sistema'>EXPORTAR</span>",
        unsafe_allow_html=True
    )
    st.markdown("")

    col_h, col_dl1, col_dl2 = st.columns([2, 2, 2])

    with col_h:
        # ── Formato Hacienda (Casilla 162 / Anexo 7) ─────────────────────────
        df_c162 = construir_df_hacienda_c162(df_fil)
        _nb_ret = cliente['nombre'].replace(' ', '_')
        st.markdown("**Formato Hacienda**")
        st.caption("Casilla 162 · Anexo 7")
        st.download_button(
            "📤 CSV Hacienda — Casilla 162",
            data=to_csv_hacienda(df_c162),
            file_name=f"C162_Ret1pct_{_nb_ret}.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )
        from utils.export_utils import _to_excel as _export_xl_ret
        _df_c162_xl = df_c162.rename(columns={
            "A": "NIT Agente Retenedor", "B": "Nombre Agente", "C": "Tipo DTE",
            "D": "Sello Recepción", "E": "UUID sin guiones",
            "F": "Base Retenida", "G": "Monto Retenido 1%", "I": "Anexo",
        })
        st.download_button(
            "📊 Excel Casilla 162 (legible)",
            data=_export_xl_ret(_df_c162_xl),
            file_name=f"C162_Ret1pct_{_nb_ret}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
        )

    with col_dl1:
        excel_bytes = generar_excel(df_fil, cliente['nombre'])
        st.markdown("**Base de Trabajo (F-14)**")
        st.caption("Excel con formato de auditoría")
        st.download_button(
            "📥 Excel Auditoría F-14",
            data=excel_bytes,
            file_name=f"Retenciones_F14_{cliente['nombre'].replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
        )
    with col_dl2:
        csv_bytes = df_fil[list(cols_existentes.keys())].to_csv(index=False).encode("utf-8")
        st.download_button(
            "📄 CSV Auditoría",
            data=csv_bytes,
            file_name=f"Retenciones_F14_{cliente['nombre'].replace(' ','_')}.csv",
            mime="text/csv",
            type="secondary",
            use_container_width=True,
        )

else:
    st.markdown("""
    <div style="text-align:center; padding: 60px 20px;">
        <p style="font-size:2.5rem; margin-bottom:8px;">📂</p>
        <h3 style="color:#6AB040 !important;">Sin documentos cargados</h3>
        <p style="color:#3A5830 !important;">
            Usa el panel lateral para cargar y procesar DTE-07 de retenciones.<br>
            El sistema extrae automáticamente base gravable y retención 1%.
        </p>
    </div>
    """, unsafe_allow_html=True)
