"""
Exportación al formato exacto F-07 — Ministerio de Hacienda El Salvador.
Columnas y nombres copiados exactamente del MVP Streamlit (pages/1_ y 2_).
POST /exportar/excel → .xlsx
"""
import io
import re
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

router = APIRouter()

_TIPOS_VALIDOS = {"ventas", "compras", "retenciones", "sujetos_excluidos"}
_TIPOS_CONTRIBUYENTES = {"03", "05", "06"}
_TIPOS_CONSUMIDOR     = {"01", "02", "10", "11"}


# ─── Utilidades ────────────────────────────────────────────────────────────

def _f(v) -> float:
    try: return round(float(v or 0), 2)
    except: return 0.0

def _s(v) -> str:
    return str(v or "")

def _clean(v) -> str:
    return re.sub(r"[-]", "", _s(v))


# ─── Compras — columnas exactas del MVP (pages/2_Extractor_DTE_Compras.py) ─

_COLS_COMPRAS_F07 = [
    "A. Fecha Emisión",         "B. Clase Documento",         "C. Tipo Documento",
    "D. Num Documento (UUID)",  "E. NIT/NRC Proveedor",       "F. Nombre Proveedor",
    "G. Compras Exentas/NS",    "H. Internac. Exentas/NS",    "I. Import. Exentas/NS",
    "J. Compras Gravadas",      "K. Internac. Grav. Bienes",  "L. Import. Grav. Bienes",
    "M. Import. Grav. Servicios","N. Crédito Fiscal (IVA)",   "O. Total Compras",
    "P. DUI Proveedor",
    "Q. Tipo Operación",        "R. Clasificación",
    "S. Sector",                "T. Tipo Costo/Gasto",        "U. Num Anexo",
]
_ANCHOS_COMPRAS = [12,2,3,38,16,45,12,12,12,12,12,12,12,12,14,10,2,2,2,2,3]

def _row_compras(r: dict, tipo_op="1", clasif="2", sector="4",
                 tipo_cg="2", periodo_feb2024=True) -> dict:
    gra = _f(r.get("gra")); exe = _f(r.get("exe")); iva = _f(r.get("iva"))
    tot = _f(r.get("tot")) or (exe + gra)
    q = tipo_op if periodo_feb2024 else "0"
    s = sector   if periodo_feb2024 else "0"
    rc = clasif  if periodo_feb2024 else "0"
    tc = tipo_cg if periodo_feb2024 else "0"
    return {
        "A. Fecha Emisión":          _s(r.get("fecha")),
        "B. Clase Documento":        "4",
        "C. Tipo Documento":         _s(r.get("tipo")),
        "D. Num Documento (UUID)":   _clean(r.get("gen_sin_guiones") or r.get("gen")),
        "E. NIT/NRC Proveedor":      _s(r.get("nit_prov")),
        "F. Nombre Proveedor":       _s(r.get("nom_prov")),
        "G. Compras Exentas/NS":     exe,
        "H. Internac. Exentas/NS":   0.0,
        "I. Import. Exentas/NS":     0.0,
        "J. Compras Gravadas":       gra,
        "K. Internac. Grav. Bienes": 0.0,
        "L. Import. Grav. Bienes":   0.0,
        "M. Import. Grav. Servicios":0.0,
        "N. Crédito Fiscal (IVA)":   iva,
        "O. Total Compras":          tot,
        "P. DUI Proveedor":          _s(r.get("dui_prov")),
        "Q. Tipo Operación":         q,
        "R. Clasificación":          rc,
        "S. Sector":                 s,
        "T. Tipo Costo/Gasto":       tc,
        "U. Num Anexo":              "3",
    }


# ─── Ventas Contribuyentes — Anexo 1 (A-T, 20 cols) ──────────────────────

_COLS_CONTRIB = [
    "A. Fecha Emisión",          "B. Clase Documento",     "C. Tipo Documento",
    "D. Num Resolución",         "E. Serie (Sello)",        "F. Num Documento (UUID)",
    "G. Control Interno",        "H. NIT/NRC Cliente",      "I. Nombre Cliente",
    "J. Ventas Exentas",         "K. Ventas No Sujetas",    "L. Ventas Gravadas",
    "M. Débito Fiscal",          "N. Vtas Cuenta Terceros", "O. Déb. Fiscal Terceros",
    "P. Total Ventas",
    "Q. DUI Cliente",            "R. Tipo Operación (Renta)",
    "S. Tipo Ingreso (Renta)",   "T. Num Anexo",
]
_ANCHOS_CONTRIB = [12,3,3,35,45,35,12,16,45,12,12,12,12,12,12,14,14,4,4,3]

def _row_contrib(r: dict, tipo_op_renta="1", tipo_ingreso_renta="3",
                 periodo_ene2025=True) -> dict:
    tor = tipo_op_renta     if periodo_ene2025 else "0"
    tir = tipo_ingreso_renta if periodo_ene2025 else "0"
    return {
        "A. Fecha Emisión":         _s(r.get("fecha")),
        "B. Clase Documento":       "4",
        "C. Tipo Documento":        _s(r.get("tipo")),
        "D. Num Resolución":        _clean(r.get("num_control")),
        "E. Serie (Sello)":         _s(r.get("sello")),
        "F. Num Documento (UUID)":  _clean(r.get("gen_sin_guiones") or r.get("gen")),
        "G. Control Interno":       "",
        "H. NIT/NRC Cliente":       _s(r.get("nit_cli")),
        "I. Nombre Cliente":        _s(r.get("nom_cli")),
        "J. Ventas Exentas":        _f(r.get("exentas")),
        "K. Ventas No Sujetas":     _f(r.get("no_sujetas")),
        "L. Ventas Gravadas":       _f(r.get("gravadas")),
        "M. Débito Fiscal":         _f(r.get("debito")),
        "N. Vtas Cuenta Terceros":  _f(r.get("terceros")),
        "O. Déb. Fiscal Terceros":  _f(r.get("deb_terc")),
        "P. Total Ventas":          _f(r.get("total")),
        "Q. DUI Cliente":           _s(r.get("dui_cli")),
        "R. Tipo Operación (Renta)":tor,
        "S. Tipo Ingreso (Renta)":  tir,
        "T. Num Anexo":             "1",
    }


# ─── Ventas Consumidor Final — Anexo 2 (A-W, 23 cols) ────────────────────

_COLS_CONSUMIDOR = [
    "A. Fecha Emisión",              "B. Clase Documento",          "C. Tipo Documento",
    "D. Num Resolución",             "E. Serie Documento",          "F. N° Control Interno DEL",
    "G. N° Control Interno AL",      "H. N° Documento DEL (UUID)",  "I. N° Documento AL (UUID)",
    "J. N° Máquina Registradora",
    "K. Ventas Exentas",             "L. Exentas No Prop.",         "M. Ventas No Sujetas",
    "N. Ventas Gravadas (c/IVA)",    "O. Export. dentro CA",        "P. Export. fuera CA",
    "Q. Export. Servicios",          "R. Vtas Zonas Francas DPA",   "S. Vtas Cuenta Terceros",
    "T. Total Ventas",
    "U. Tipo Operación (Renta)",     "V. Tipo Ingreso (Renta)",     "W. Num Anexo",
]
_ANCHOS_CONSUMIDOR = [12,3,3,8,8,8,8,35,35,16,12,12,12,14,12,12,12,12,12,14,4,4,3]

def _row_consumidor(r: dict, tipo_op_renta="1", tipo_ingreso_renta="3",
                    periodo_ene2025=True) -> dict:
    gen_sg = _clean(r.get("gen_sin_guiones") or r.get("gen"))
    tipo   = _s(r.get("tipo"))
    tor = tipo_op_renta     if periodo_ene2025 else "0"
    tir = tipo_ingreso_renta if periodo_ene2025 else "0"
    return {
        "A. Fecha Emisión":             _s(r.get("fecha")),
        "B. Clase Documento":           "4",
        "C. Tipo Documento":            tipo,
        "D. Num Resolución":            "N/A",
        "E. Serie Documento":           "N/A",
        "F. N° Control Interno DEL":    "N/A",
        "G. N° Control Interno AL":     "N/A",
        "H. N° Documento DEL (UUID)":   gen_sg,
        "I. N° Documento AL (UUID)":    gen_sg,
        "J. N° Máquina Registradora":  _s(r.get("num_maq")) if tipo == "10" else "",
        "K. Ventas Exentas":            _f(r.get("exentas")),
        "L. Exentas No Prop.":          0.0,
        "M. Ventas No Sujetas":         _f(r.get("no_sujetas")),
        "N. Ventas Gravadas (c/IVA)":   _f(r.get("gravadas")),
        "O. Export. dentro CA":         0.0,
        "P. Export. fuera CA":          0.0,
        "Q. Export. Servicios":         0.0,
        "R. Vtas Zonas Francas DPA":    0.0,
        "S. Vtas Cuenta Terceros":      _f(r.get("terceros")),
        "T. Total Ventas":              _f(r.get("total")),
        "U. Tipo Operación (Renta)":    tor,
        "V. Tipo Ingreso (Renta)":      tir,
        "W. Num Anexo":                 "2",
    }


# ─── Retenciones — Anexo 7 (A-I, 9 cols) ──────────────────────────────────

_COLS_RETENCION = [
    "A. NIT Agente",        "B. Fecha Emisión",      "C. Tipo Documento",
    "D. Serie Documento",   "E. Num Documento",
    "F. Monto Sujeto",      "G. Monto Retención 1%",
    "H. DUI Agente",        "I. Num Anexo",
]
_ANCHOS_RETENCION = [16,12,3,40,36,14,14,12,3]

def _row_retencion(r: dict) -> dict:
    return {
        "A. NIT Agente":       _s(r.get("nit_prov")),
        "B. Fecha Emisión":    _s(r.get("fecha")),
        "C. Tipo Documento":   _s(r.get("tipo") or "07"),
        "D. Serie Documento":  _s(r.get("sello")),
        "E. Num Documento":    _clean(r.get("gen")),
        "F. Monto Sujeto":     _f(r.get("base")),
        "G. Monto Retención 1%": _f(r.get("ret")),
        "H. DUI Agente":       _s(r.get("dui_agente")),
        "I. Num Anexo":        "7",
    }


# ─── Sujetos Excluidos — Anexo 5 (A-M, 13 cols) ───────────────────────────

_COLS_SUJETOS = [
    "A. Tipo Doc Identidad",  "B. Num Identificación",  "C. Nombre Sujeto Excluido",
    "D. Fecha Emisión",       "E. Serie Documento",      "F. Num Documento",
    "G. Monto Operación",     "H. Retención IVA 13%",
    "I. Tipo Operación",      "J. Clasificación",        "K. Sector",
    "L. Tipo Costo/Gasto",    "M. Num Anexo",
]
_ANCHOS_SUJETOS = [3,16,45,12,40,36,14,14,3,3,3,3,3]

def _row_sujeto(r: dict) -> dict:
    id_raw = re.sub(r"[-\s]", "", _s(r.get("id_sujeto")))
    tipo_id = "1" if len(id_raw)==14 and id_raw.isdigit() else ("2" if len(id_raw)==9 and id_raw.isdigit() else "3")
    return {
        "A. Tipo Doc Identidad":    tipo_id,
        "B. Num Identificación":    id_raw,
        "C. Nombre Sujeto Excluido":_s(r.get("nom_sujeto")),
        "D. Fecha Emisión":         _s(r.get("fecha")),
        "E. Serie Documento":       _s(r.get("sello")),
        "F. Num Documento":         _clean(r.get("gen") or r.get("num_control")),
        "G. Monto Operación":       _f(r.get("base")),
        "H. Retención IVA 13%":     _f(r.get("ret")),
        "I. Tipo Operación":        _s(r.get("tipo_operacion") or "0"),
        "J. Clasificación":         _s(r.get("clasificacion") or "0"),
        "K. Sector":                _s(r.get("sector") or "0"),
        "L. Tipo Costo/Gasto":      _s(r.get("tipo_costo_gasto") or "0"),
        "M. Num Anexo":             "5",
    }


# ─── Excel builder ──────────────────────────────────────────────────────────

def _build_xlsx(sheets: dict) -> bytes:
    """
    sheets: {nombre_hoja: (columnas, filas, anchos, col_num_ini, col_num_fin)}
    Primera fila = nombres de columnas (header=True).
    Sin filas de título. Formato numérico #,##0.00 en columnas monetarias.
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas no instalado")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre, (cols, filas, anchos, num_ini, num_fin) in sheets.items():
            df = pd.DataFrame(filas, columns=cols) if filas else pd.DataFrame(columns=cols)
            # header=True → primera fila = nombres de columnas exactos del Anexo
            df.to_excel(writer, index=False, sheet_name=nombre[:31])

            ws = writer.sheets[nombre[:31]]
            # Anchos exactos del MVP
            for idx, ancho in enumerate(anchos, 1):
                try:
                    ws.column_dimensions[ws.cell(1, idx).column_letter].width = ancho
                except Exception:
                    pass
            # Formato numérico en columnas monetarias
            if num_ini and num_fin:
                for fila in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         min_col=num_ini, max_col=num_fin):
                    for celda in fila:
                        if isinstance(celda.value, (int, float)):
                            celda.number_format = '#,##0.00'
            # Estilo encabezado
            try:
                from openpyxl.styles import Font, PatternFill
                fill = PatternFill("solid", fgColor="1C2333")
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = fill
            except Exception:
                pass

    return buf.getvalue()


# ─── Request model ──────────────────────────────────────────────────────────

class ExportarRequest(BaseModel):
    tipo: str
    declarante_id: str
    periodo: Optional[str] = None
    registros: List[Dict[str, Any]]
    # Compras Q-T
    tipo_op:        str  = "1"
    clasif:         str  = "2"
    sector:         str  = "4"
    tipo_cg:        str  = "2"
    periodo_feb2024: bool = True
    # Ventas R-S / U-V
    tipo_op_renta:      str  = "1"
    tipo_ingreso_renta: str  = "3"
    periodo_ene2025:    bool = True


# ─── Endpoint ───────────────────────────────────────────────────────────────

@router.post("/excel")
async def exportar_excel(body: ExportarRequest):
    """
    Genera xlsx F-07 con columnas exactas del Anexo.
    Primera fila = nombres de columnas. Sin filas de título.
    """
    if body.tipo not in _TIPOS_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo inválido. Debe ser uno de: {', '.join(_TIPOS_VALIDOS)}",
        )

    periodo_str = body.periodo or "sin_periodo"
    sheets: dict = {}

    if body.tipo == "ventas":
        contrib = [_row_contrib(r, body.tipo_op_renta, body.tipo_ingreso_renta, body.periodo_ene2025)
                   for r in body.registros if _s(r.get("tipo")) in _TIPOS_CONTRIBUYENTES
                   or _s(r.get("tipo")) not in _TIPOS_CONTRIBUYENTES | _TIPOS_CONSUMIDOR]
        consumidor = [_row_consumidor(r, body.tipo_op_renta, body.tipo_ingreso_renta, body.periodo_ene2025)
                      for r in body.registros if _s(r.get("tipo")) in _TIPOS_CONSUMIDOR]
        sheets["Ventas_Contribuyentes"] = (_COLS_CONTRIB,    contrib,   _ANCHOS_CONTRIB,    10, 16)
        sheets["Ventas_Consumidor"]     = (_COLS_CONSUMIDOR, consumidor,_ANCHOS_CONSUMIDOR, 11, 20)

    elif body.tipo == "compras":
        filas = [_row_compras(r, body.tipo_op, body.clasif, body.sector,
                              body.tipo_cg, body.periodo_feb2024)
                 for r in body.registros]
        sheets["Compras_F07"] = (_COLS_COMPRAS_F07, filas, _ANCHOS_COMPRAS, 7, 15)

    elif body.tipo == "retenciones":
        filas = [_row_retencion(r) for r in body.registros]
        sheets["Retenciones_Anexo7"] = (_COLS_RETENCION, filas, _ANCHOS_RETENCION, 6, 7)

    elif body.tipo == "sujetos_excluidos":
        filas = [_row_sujeto(r) for r in body.registros]
        sheets["SujetosExcluidos_Anexo5"] = (_COLS_SUJETOS, filas, _ANCHOS_SUJETOS, 7, 8)

    nombre_tipo = {
        "ventas": "Ventas", "compras": "Compras",
        "retenciones": "Retenciones", "sujetos_excluidos": "SujetosExcluidos",
    }[body.tipo]
    filename = f"F07_{nombre_tipo}_{body.declarante_id}_{periodo_str}.xlsx"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    xlsx_bytes = _build_xlsx(sheets)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
